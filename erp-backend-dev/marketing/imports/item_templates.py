import openpyxl
from marketing.models import Item, ItemAttribute

headers={
    'item_description': 'Item Description',
    'item_reference_code': 'Item reference/code',
    'composition': 'Composition',
    'gsm': 'GSM',
    'placement': 'Placement',
    'supplier': 'Supplier',
    'consumption': 'Consumption',
    'size': 'Size'
}
main_headers={
    'fabric': 'Fabric',
    'trim': 'Sewing Trims'
}
special_headers=['fabric', 'trim']
def get_key_by_value(value,dictionary):
    for key in dictionary:
        if dictionary[key] == value:
            return key

def create_item_attribute(item_data):
    name = item_data['name']
    try:
        item = Item.objects.get_or_create(name=name)[0]
        for attribute_type in ['fabric', 'trim']:
            for attribute_data in item_data[attribute_type]:
                item_attribute = ItemAttribute.objects.get_or_create(item=item, placement=attribute_data['placement'], type=attribute_type)[0]
        return {'id': item.id, 'name': item.name, 'import': 'Completed'}
    except:
        return "Import Error"

def read_templete(templates_file):
    excel_doc = openpyxl.load_workbook(filename=templates_file)
    data_set = []
    for sheet in excel_doc.worksheets:
        fabric_cells=[]
        for row in range(1, sheet.max_row):
            for col in range(1, sheet.max_column):
                if sheet.cell(row=row, column=col).value == main_headers['fabric']:
                    fabric_cells.append(sheet.cell(row=row, column=col))
        for fabric_cell in fabric_cells:
            item=""
            for item_row in range(fabric_cell.row-1, 1, -1):
                cell = sheet.cell(row=item_row, column=fabric_cell.column)
                if cell.value != None:
                    item=cell.value
                    break
            if item:
                items={}
                items['name'] = item
                def read_item_detail(main_cell):
                    for row in range(main_cell.row, sheet.max_row):
                        bottom = sheet.cell(row=row, column=main_cell.column).border.bottom.style
                        if bottom != None:
                            last_row = row
                            break
                    
                    main_row = main_cell.row
                    while True:
                        type = get_key_by_value(sheet.cell(row=main_row, column=main_cell.column).value, main_headers)
                        if type:
                            type_detail = []
                            special = type in special_headers
                            for row in range(main_row+2, last_row):
                                if special:
                                    if sheet.cell(row=row+1, column=main_cell.column).value == None:
                                        main_row = row+1
                                        break
                                else:
                                    if sheet.cell(row=row, column=main_cell.column).value == None:
                                        main_row = row
                                        break
                                temp_detail={}
                                for col in range(main_cell.column, sheet.max_column):
                                    header = sheet.cell(row=main_row+1, column=col).value
                                    key = get_key_by_value(header, headers)
                                    if key != None:
                                        value = sheet.cell(row=row, column=col).value
                                        if value != None:
                                            temp_detail[key] = value
                                    
                                type_detail.append(temp_detail)
                            items[type] = type_detail
                        main_row = main_row + 1
                        if main_row == last_row:
                            break
                    return items

                
                return_cell = read_item_detail(fabric_cell)
                # data_set.append(return_cell)
                # print(return_cell)
                imported = create_item_attribute(return_cell)
                data_set.append(imported)
    excel_doc.close()
    return data_set