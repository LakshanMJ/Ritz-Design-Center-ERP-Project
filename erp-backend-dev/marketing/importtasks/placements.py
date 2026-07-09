import openpyxl
from marketing.models import Item, ItemAttribute
from django.conf import settings
import os

def iterate_through_placement_data(imported_sheet):
    product_col = 0
    product_description_col = 0
    rm_type_col = 0
    placement_col = 0

    max_row = 0
    max_col = 6
    for row in imported_sheet:
        if not all([cell.value is None for cell in row]):
            max_row += 1

    for c in range(1,max_col):
        if imported_sheet.cell(1,c).value == "Product":
            product_col = c
        elif imported_sheet.cell(1,c).value == "Product Description":
            product_description_col = c
        elif imported_sheet.cell(1,c).value == "RM Type":
            rm_type_col = c
        elif imported_sheet.cell(1,c).value == "Placement":
            placement_col = c


    for r in range(2,max_row+1):
            product_temp = imported_sheet.cell(row=r, column=product_col).value
            product_description_temp = imported_sheet.cell(row=r, column=product_description_col).value
            rm_type_temp = imported_sheet.cell(row=r, column=rm_type_col).value
            placement = imported_sheet.cell(row=r, column=placement_col).value
            # print(product_temp, placement)

            if product_temp != None:
                product = product_temp
            if product_description_temp != None:
                product_description = product_description_temp
            if rm_type_temp != None:
                rm_type = rm_type_temp


def saving_to_database(product, product_description, rm_type, placement):
            item_name = product+" "+product_description
            
            Item_dictionary = {                
                'name': item_name,
                'active' : True
            }
            item_instance = Item.objects.get(id=1)
            
            ItemAttribute_dictionary = {
                'item_id': item_instance.id,
                'type': rm_type,
                'placement': placement,
                'active': True
            }

            if product != None and product_description == None:
                name = Item.objects.get_or_create(**Item_dictionary)
            type_and_placement = ItemAttribute.objects.get_or_create(**ItemAttribute_dictionary)


def import_placement_data():
    base_dir = settings.BASE_DIR
    file_path = "marketing/dataimports/data_files/ProductInfo.xlsx"

    excel_file_path = os.path.join(base_dir, file_path)
    excel_doc = openpyxl.load_workbook(filename=excel_file_path)
    imported_sheet = excel_doc.active
    iterate_through_placement_data(imported_sheet)





