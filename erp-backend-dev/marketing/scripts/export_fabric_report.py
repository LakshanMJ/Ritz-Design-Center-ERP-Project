from io import BytesIO

from django.http import HttpResponse
from materials.models import FABRIC_TRIM_TYPES
from shared.models import InHouseMaterial
import openpyxl
from openpyxl.utils import get_column_letter
from supplier_po.models import SupplierPOGRNMaterialDetail

STYLE_HEADER = 'Style'
ORDER_NO_HEADER = 'Order No'
ITEM_DESCRIPTION_HEADER = 'Item Description'
COLOR_HEADER = 'Colour'
ROLL_NO_HEADER = 'Roll No'
BATCH_NO_HEADER = 'Batch No'
BARCODE_HEADER = 'Barcode'
SHADE_LOT_HEADER = 'Shade Lot'
WIDTH_HEADER = 'Width'
UNIT_HEADER = 'Unit'
QTY_HEADER = 'Qty'
REMARKS_HEADER = 'Remarks'
MARKER_TYPE_HEADER = 'Marker Type'

class FabricReportExporter:

    po_club = None

    def __init__(self, po_club):
        self.po_club = po_club
    
    def get_fabric_inhouse_materials(self):
        return InHouseMaterial.objects.filter(grn_material_detail__supplier_po_grn_material__supplier_po_grn__supplier_po__general_po_supplier__general_po__po_club=self.po_club,
                                                grn_material_detail__supplier_po_grn_material__grn_material__customer_brand_material__material_detail__generic_material__user_material__category=FABRIC_TRIM_TYPES
                                              )

    
    def export_excel_report(self):
        inhouse_materials = self.get_fabric_inhouse_materials()
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Fabric Report"
        headers = [STYLE_HEADER, ORDER_NO_HEADER, ITEM_DESCRIPTION_HEADER, COLOR_HEADER, ROLL_NO_HEADER, BATCH_NO_HEADER, BARCODE_HEADER, SHADE_LOT_HEADER, WIDTH_HEADER, UNIT_HEADER, QTY_HEADER, REMARKS_HEADER, MARKER_TYPE_HEADER]
        row = 1
        for col, header in enumerate(headers, start=1):
            sheet[f"{get_column_letter(col)}{row}"] = header
        row += 1
        for inhouse_material in inhouse_materials:
            user_defined_material_data = inhouse_material.grn_material_detail.supplier_po_grn_material.grn_material.customer_brand_material.material_detail.user_defined_material_data
            color = user_defined_material_data.get('fabric_color', '')
            shade_category = inhouse_material.grn_material_detail.shade_category
            marker_type = 'Close Marker' if shade_category == SupplierPOGRNMaterialDetail.SHADE_CATEGORY_WITHIN_THE_ROLL_SHADING else 'Normal Marker'
            row_data = {
                STYLE_HEADER: self.po_club.style_number,
                ORDER_NO_HEADER: self.po_club.display_number,
                ITEM_DESCRIPTION_HEADER: inhouse_material.grn_material_detail.supplier_po_grn_material.grn_material.customer_brand_material.verbose_reference_code,
                COLOR_HEADER: color,
                ROLL_NO_HEADER: inhouse_material.grn_material_detail.fabricgrndetail.pack_number,
                BATCH_NO_HEADER: inhouse_material.grn_material_detail.batch_number.batch_number,
                BARCODE_HEADER: inhouse_material.barcode,
                SHADE_LOT_HEADER: inhouse_material.grn_material_detail.shade.shade,
                WIDTH_HEADER: inhouse_material.grn_material_detail.fabricgrndetail.actual_width.actual_width,
                UNIT_HEADER: 'Mtrs',#inhouse_material.get_available_quantity_units_display(),
                QTY_HEADER: inhouse_material.get_allocated_quantity().get('quantity', 0),
                REMARKS_HEADER: inhouse_material.grn_material_detail.fabricgrndetail.remarks,
                MARKER_TYPE_HEADER: marker_type
            }
            if row_data.get(QTY_HEADER, 0) > 0:
                for col, header in enumerate(headers, start=1):
                    sheet[f"{get_column_letter(col)}{row}"] = row_data.get(header, None)
                row += 1
        excel_file = BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=purchase_order_report.xlsx'
        workbook.close()
        return response