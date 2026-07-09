from openpyxl import load_workbook, Workbook
import os
from marketing.utils.aws_utils import handle_file_read
from marketing.models import POFabricMarker, POCADMarkerUpload, POFabricMarkerPlacement, FabricWidth, POMarkerPoint
from materials.fieldmetadata.measuring_unit_helpers import MaterialUnitHelper
from shared.utils import convert_quantity_to_unit

class PlacementAreaDetailFileReader:
    area_file = None
    marker_sizes = []
    marker_placements = []
    file_type = '.xlsx'

    SIZE_HEADER = 'size'
    PLACEMENT_HEADER = 'placement'
    QUANTITY_HEADER = 'quantity'
    AREA_HEADER = 'area'
    # SIZE_ID_HEADER = 'size_id'
    # PLACEMENT_ID_HEADER = 'placement_id'


    SIZE_HEADER_DISPLAY = 'Size'
    PLACEMENT_HEADER_DISPLAY = 'Piece Name'
    QUANTITY_HEADER_DISPLAY = 'Quantity'
    AREA_HEADER_DISPLAY = 'Area'
    # SIZE_ID_HEADER_DISPLAY = 'Size ID'
    # PLACEMENT_ID_HEADER_DISPLAY = 'Placement ID'


    HEADERS = (
        (SIZE_HEADER, SIZE_HEADER_DISPLAY),
        (PLACEMENT_HEADER, PLACEMENT_HEADER_DISPLAY),
        (QUANTITY_HEADER, QUANTITY_HEADER_DISPLAY),
        (AREA_HEADER, AREA_HEADER_DISPLAY),
        # (SIZE_ID_HEADER, SIZE_ID_HEADER_DISPLAY),
        # (PLACEMENT_ID_HEADER, PLACEMENT_ID_HEADER_DISPLAY)
    )

    def get_mapped_marker_size(self, size):
        mapped_size = None
        for marker_size in self.marker_sizes:
            if marker_size == size.name or marker_size == size.abbreviation:
                mapped_size = marker_size
                break
        return mapped_size
    
    def get_mapped_marker_placements(self, placement):
        mapped_placements = []
        cad_marker_placements = placement.cadmarkerplacement_set.all().distinct('placement_name')
        for cad_marker_placement in cad_marker_placements:
            if cad_marker_placement.placement_name in self.marker_placements:
                mapped_placements.append(cad_marker_placement.placement_name)
        return mapped_placements

    def file_open(self, file_path):
        if file_path.endswith(self.file_type):
            file = handle_file_read(file_path)
            self.area_file = load_workbook(file)

    def __init__(self, file_path):
        self.file_open(file_path=file_path)
    
    def is_file_open(self):
        return type(self.area_file) == Workbook
    
    def file_close(self):
        if self.is_file_open():
            self.area_file.close()
            self.area_file = None
    
    def have_all_headers(self, header_details):
        return all(header[0] in header_details for header in self.HEADERS)
    
    def header_validate(self, header_details):
        errors = []
        header_row = None
        for header, display in self.HEADERS:
            if not header in header_details:
                errors.append(display + ' column cannot find')
        header_rows = list(set([header_details[header].row for header in header_details]))
        if not len(header_rows) == 1:
            errors.append("Headers are not in same row")
        else:
            header_row = header_rows[0]
        return not errors == [], errors, header_row
    
    def get_file_headers(self):
        header_details = {}
        errors = []
        has_errors = False
        
        if self.is_file_open():
            for row in self.area_file.active.iter_rows():
                for cell in row:
                    cell_value = cell.value
                    for header, diaplay in self.HEADERS:
                        if diaplay == cell_value:
                            header_details[header] = cell
                            if self.have_all_headers(header_details):
                                break
                    else:
                        continue
                    break
                else:
                    continue
                break
        else:
            errors.append("File is not opened")
            has_errors = True
        has_validate_errors, validate_errors, header_row = self.header_validate(header_details)
        if has_validate_errors:
            has_errors = True
            errors.extend(validate_errors)
        return has_errors, errors, header_details, header_row
    
    def file_read(self):
        has_errors, errors, header_details, header_row = self.get_file_headers()
        max_row = self.area_file.active.max_row
        sizes = []
        placements = []
        area_data = {}
        if not has_errors:
            sheet = self.area_file.active
            placement = None
            area = None
            size_id = None
            placement_id = None
            for row_number in range(header_row + 1, max_row):
                size = sheet.cell(row=row_number, column=header_details[self.SIZE_HEADER].column).value
                # size_id_temp = sheet.cell(row=row_number, column=header_details[self.SIZE_ID_HEADER].column).value
                # if not size_id_temp == None:
                #     size_id = size_id_temp
                placement_temp = sheet.cell(row=row_number, column=header_details[self.PLACEMENT_HEADER].column).value
                # placement_id_temp = sheet.cell(row=row_number, column=header_details[self.PLACEMENT_ID_HEADER].column).value
                # if not placement_id_temp == None:
                #     placement_id = placement_id_temp
                if not placement_temp == None:
                    placement = placement_temp
                quantity = sheet.cell(row=row_number, column=header_details[self.QUANTITY_HEADER].column).value
                area_temp = sheet.cell(row=row_number, column=header_details[self.AREA_HEADER].column).value
                if not area_temp == None:
                    area = area_temp
                if all([size, placement, quantity, area]):
                    size_data = size
                    if not size_data in sizes:
                        sizes.append(size_data)
                        area_data[size] = {}
                    if quantity > 1:
                        for index in range(1, quantity + 1):
                            placement_data = placement + '_' + str(index)
                            if not placement_data in placements:
                                placements.append(placement_data)
                            area_data[size][placement_data] = {
                                'area': area
                            }
                    else:
                        placement_data = placement
                        if not placement_data in placements:
                            placements.append(placement_data)
                        area_data[size][placement_data] = {
                            'area': area
                        }
        self.marker_sizes = sizes
        self.marker_placements = placements
        return {
            'sizes': sizes,
            'placements': placements,
            'area_data': area_data
        }


# placement_area_detail_file_reader = PlacementAreaDetailFileReader('ST400-2nd BULK PATTERN-27-04-22-BOOKING-NOVEMBER-1138648-BLK TRD SOLID-1A-68 IN_ Area.xlsx')
# placement_area_detail_file_reader.file_read()
# placement_area_detail_file_reader.file_close()


class MiniMarkerFileReader:
    file_type = '.xlsx'
    mini_marker_file = None

    #general information
    MARKER_NAME_HEADER = 'marker_name'
    MARKER_WIDTH_HEADER = 'marker_width'
    MARKER_LENGTH_HEADER = 'marker_length'
    MARKER_CUT_LENGTH_HEADER = 'marker_cut_length'

    MARKER_NAME_HEADER_DISPLAY = 'Marker Name'
    MARKER_WIDTH_HEADER_DISPLAY = 'Marker Width'
    MARKER_LENGTH_HEADER_DISPLAY = 'Marker Length'
    MARKER_CUT_LENGTH_HEADER_DISPLAY = 'Cut.Length'

    #size ratio
    SIZE_HEADER = 'size'
    SIZE_RATIO_HEADER = 'size_ratio'

    SIZE_HEADER_DISPLAY = 'Size'
    SIZE_RATIO_HEADER_DISPLAY = 'Order'

    #marker points
    CUT_POINT_HEADER = 'cut_point'
    BACK_POINT_HEADER = 'back_point'

    CUT_POINT_HEADER_DISPLAY = 'Cut Position'
    BACK_POINT_HEADER_DISPLAY = 'Back Point'

    ALL_HEADERS = (
        (MARKER_NAME_HEADER, MARKER_NAME_HEADER_DISPLAY),
        (MARKER_WIDTH_HEADER, MARKER_WIDTH_HEADER_DISPLAY),
        (MARKER_LENGTH_HEADER, MARKER_LENGTH_HEADER_DISPLAY),
        (MARKER_CUT_LENGTH_HEADER, MARKER_CUT_LENGTH_HEADER_DISPLAY),
        (SIZE_HEADER, SIZE_HEADER_DISPLAY),
        (SIZE_RATIO_HEADER, SIZE_RATIO_HEADER_DISPLAY),
        (CUT_POINT_HEADER, CUT_POINT_HEADER_DISPLAY),
        (BACK_POINT_HEADER, BACK_POINT_HEADER_DISPLAY)
    )

    SIZE_RATIO_DATA = 'size_ratio_data'
    MARKER_POINT_DATA = 'marker_point_data'

    TABLE_HEADERS = {
        SIZE_RATIO_DATA: [SIZE_HEADER, SIZE_RATIO_HEADER],
        MARKER_POINT_DATA: [CUT_POINT_HEADER, BACK_POINT_HEADER]
    }

    def is_valid(self, mini_marker_file_data):
        valid = True
        validation_data = {
            self.MARKER_NAME_HEADER: None,
            self.MARKER_WIDTH_HEADER: None,
            self.MARKER_LENGTH_HEADER: None,
            self.SIZE_RATIO_DATA: []
        }

        for validation_field in validation_data:
            if mini_marker_file_data[validation_field] == validation_data[validation_field]:
                valid = False

        return valid

    def file_open(self, file_path):
        if file_path.endswith(self.file_type):
            file = handle_file_read(file_path)
            self.mini_marker_file = load_workbook(file)

    def __init__(self, file_path):
        self.file_open(file_path=file_path)
    
    def is_file_open(self):
        return type(self.mini_marker_file) == Workbook
    
    def file_close(self):
        if self.is_file_open():
            self.mini_marker_file.close()
            self.mini_marker_file = None

    def have_all_headers(self, header_details):
        return all(header[0] in header_details for header in self.ALL_HEADERS)
    
    def header_validate(self, header_details):
        errors = []
        row_data = {}
        for header, display in self.ALL_HEADERS:
            if not header in header_details:
                errors.append(display + ' column cannot find')
        if self.have_all_headers(header_details):
            if not header_details[self.SIZE_HEADER].row == header_details[self.SIZE_RATIO_HEADER].row:
                errors.append("Size & Order headers are not in same row")
            else:
                row_data[self.SIZE_RATIO_DATA] = header_details[self.SIZE_HEADER].row
            if not header_details[self.CUT_POINT_HEADER].row == header_details[self.BACK_POINT_HEADER].row:
                errors.append("Cut point & Back point headers are not in same row")
            else:
                row_data[self.MARKER_POINT_DATA] = header_details[self.CUT_POINT_HEADER].row
        return not errors == [], errors, row_data
    
    def get_file_headers(self):
        header_details = {}
        errors = []
        has_errors = False
        
        if self.is_file_open():
            for row in self.mini_marker_file.active.iter_rows():
                for cell in row:
                    cell_value = cell.value
                    for header, diaplay in self.ALL_HEADERS:
                        if diaplay == cell_value:
                            header_details[header] = cell
                            if self.have_all_headers(header_details):
                                break
                    else:
                        continue
                    break
                else:
                    continue
                break
        else:
            errors.append("File is not opened")
            has_errors = True
        has_validate_errors, validate_errors, row_data = self.header_validate(header_details)
        if has_validate_errors:
            has_errors = True
            errors.extend(validate_errors)
        return has_errors, errors, header_details, row_data
    
    def get_marker_length(self, marker_length_value):
        marker_length = 0.00
        for data in marker_length_value.split(' '):
            if data[len(data)-2: len(data)] == 'cm':
                marker_length += float(data[:len(data)-2])#+3
            elif data[len(data)-1:len(data)] == 'm':
                marker_length += float(data[:len(data)-1])*100
        if marker_length>0:
            marker_length = marker_length/100
            marker_length = round(marker_length, 4)
        return marker_length
    
    def get_marker_width(self, marker_width_value):
        marker_width = 0.00
        marker_width_value = float(marker_width_value.split(' ')[0])
        marker_width = convert_quantity_to_unit(MaterialUnitHelper.METERS_UNIT, marker_width_value, MaterialUnitHelper.CENTIMETERS_UNIT)
        return marker_width
    
    def file_read(self):
        has_errors, errors, header_details, row_data = self.get_file_headers()
        return_data = {
            self.MARKER_NAME_HEADER: None,
            self.MARKER_WIDTH_HEADER: None,
            self.MARKER_LENGTH_HEADER: None,
            self.MARKER_CUT_LENGTH_HEADER: None,
            self.SIZE_RATIO_DATA: [],
            self.MARKER_POINT_DATA: []
        }
        if not has_errors:
            sheet = self.mini_marker_file.active
            max_row = sheet.max_row
            for key in return_data:
                if key in self.TABLE_HEADERS:
                    for row_number in range(row_data[key]+1, max_row):
                        temp = {}
                        for header in self.TABLE_HEADERS[key]:
                            temp[header] = sheet.cell(row=row_number, column=header_details[header].column).value
                            if temp[header] == 'All':
                                # temp[header] = None
                                continue
                            if temp[header] == None:
                                break
                        else:
                            return_data[key].append(temp)
                            continue
                        break
                else:
                    return_data[key] = sheet.cell(row=header_details[key].row, column=header_details[key].column+1).value
                    if key == self.MARKER_LENGTH_HEADER:
                        return_data[key] = self.get_marker_length(return_data[key])
                    elif key == self.MARKER_WIDTH_HEADER:
                        return_data[key] = self.get_marker_width(return_data[key])
        return return_data
        


# mini_marker_file_rader = MiniMarkerFileReader('ST400-2nd BULK PATTERN-27-04-22-BOOKING-NOVEMBER-1138648-BLK TRD SOLID-1A-68 IN.xlsx')
# mini_marker_file_rader.file_read()
# mini_marker_file_rader.file_close()

class POFabricMarkerCreator:
    area_file_reader = None
    mini_marker_file_reader = None
    area_file = None
    mini_marker_file = None
    mapped_sizes = []
    mapped_placements = []
    def __init__(self, area_file, mini_marker_file, mapped_sizes, mapped_placements):
        self.area_file = area_file
        self.mini_marker_file = mini_marker_file
        self.area_file_reader = PlacementAreaDetailFileReader(area_file.file_path)
        self.mini_marker_file_reader = MiniMarkerFileReader(mini_marker_file.file_path)
        self.mapped_sizes = mapped_sizes
        self.mapped_placements = mapped_placements

    def close(self):
        if type(self.area_file_reader) == PlacementAreaDetailFileReader:
            self.area_file_reader.file_close()
        if type(self.mini_marker_file_reader) == MiniMarkerFileReader:
            self.mini_marker_file_reader.file_close()

    def get_po_cad_marker_upload(self):
        po_cad_marker_upload, created = POCADMarkerUpload.objects.get_or_create(
            area_file=self.area_file,
            mini_marker_file=self.mini_marker_file
        )
        return po_cad_marker_upload

    def create_marker_points(self, po_fabric_marker, marker_point_data):
        has_errors = False
        for marker_point in marker_point_data:
            POMarkerPoint.objects.create(
                po_marker=po_fabric_marker,
                cut_point=marker_point['cut_point'],
                back_point=marker_point['back_point'],
                cut_point_unit=MaterialUnitHelper.CENTIMETERS_UNIT,
                back_point_unit=MaterialUnitHelper.CENTIMETERS_UNIT
            )
        return not has_errors

    def create_marker(self, actual_po_club, material, placements):
        has_errors = False
        mini_marker_file_data = self.mini_marker_file_reader.file_read()
        # print(mini_marker_file_data)
        parent_marker = None
        try:
            parent_marker = POFabricMarker.objects.get(
                actual_club=actual_po_club,
                po_material=material,
                parent_marker=None,
            )
        except:
            pass
        po_fabric_marker = POFabricMarker.objects.create(
            actual_club=actual_po_club,
            po_material=material,
            parent_marker=parent_marker,
            marker_upload=self.get_po_cad_marker_upload(),
            marker_name=mini_marker_file_data['marker_name'],
            marker_length=mini_marker_file_data['marker_length'],
            marker_length_unit=MaterialUnitHelper.METERS_UNIT,
            marker_type=POFabricMarker.ITEM_LEVEL_MARKER,
        )
        total_placement_count = 0
        for size_ratio in mini_marker_file_data['size_ratio_data']:
            size_data = {}
            for mapped_size in self.mapped_sizes:
                if mapped_size['marker_size'] == size_ratio['size']:
                    size_data = mapped_size
                    break
            if not size_data == {}:
                size_placements = placements.filter(po_pack_item__po_pack__po_size__order_size__size__id=size_data['id'])
                if size_placements.exists():
                    for size_placement in size_placements:
                        total_placement_count += size_ratio['size_ratio']
                        po_fabric_marker_placement = POFabricMarkerPlacement.objects.create(
                            marker=po_fabric_marker,
                            placement=size_placement,
                            ratio=size_ratio['size_ratio']
                        )
                        po_fabric_marker_placement.save_cad_marker_placement_area()
        po_fabric_marker.consumption_ratio = int((po_fabric_marker.marker_length/total_placement_count)*10000)/10000
        fabric_widths = FabricWidth.objects.filter(actual_po_club=actual_po_club, customer_brand_material = material)
        for fabric_width in fabric_widths:
            if mini_marker_file_data['marker_width']['quantity'] == int(fabric_width.get_normalized_width()['quantity']*10000)/10000:
                po_fabric_marker.width = fabric_width
                break
        po_fabric_marker.save()
        self.create_marker_points(po_fabric_marker, mini_marker_file_data['marker_point_data'])
        self.close()
        return not has_errors