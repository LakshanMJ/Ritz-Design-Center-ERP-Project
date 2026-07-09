import openpyxl
from datetime import datetime
from marketing.utils.aws_utils import handle_file_read
from marketing.mixins.po_processor_mixins import POProcessorMixin


class LIDLPOProcessor(POProcessorMixin):
    _color_text = ['Color', 'color', ]
    _total_ratio_text = ['Total Ratio', 'total ratio']
    _total_qty_text = ['Total Qty', 'total qty']
    _delivery_date_text = ['Delivery Date', 'delivery date']
    _style_text = ['Style', 'style']
    _style = None
    _delivery_date = None

    def __init__(self, excel_file_path, customer_id, attachment_id):
        super().__init__()
        self.excel_file_path = excel_file_path
        self.customer_id = customer_id
        self.attachment_id = attachment_id

    def santize_value(self, value):
        return str(value).lower().strip()
    
    def get_delivery_date(self, sheet, current_row_index, current_column_index):
        delivery_date = None
        cell_value = self.santize_value(sheet.cell(row=current_row_index, column=current_column_index).value)
        if cell_value in self._delivery_date_text:
            self._delivery_date = sheet.cell(row=current_row_index, column=current_column_index + 1).value
    
    def get_sheet_style(self, sheet, current_row_index, current_column_index):
        style = None
        cell_value = self.santize_value(sheet.cell(row=current_row_index, column=current_column_index).value)
        if cell_value in self._style_text:
            self._style = sheet.cell(row=current_row_index, column=current_column_index + 1).value

    def is_po_countries_block(self, sheet, current_row_index, current_column_index):
        has_errors = False
        is_header_row = False
        color_cell_value = self.santize_value(sheet.cell(row=current_row_index, column=current_column_index).value)
        total_ratio_cell_value = self.santize_value(sheet.cell(row=current_row_index, column=current_column_index + 6).value)
        total_qty_cell_value = self.santize_value(sheet.cell(row=current_row_index, column=current_column_index + 7).value)

        if color_cell_value in self._color_text and total_ratio_cell_value in self._total_ratio_text and total_qty_cell_value in self._total_qty_text:
            is_header_row = True
            #print(color_cell_value, total_ratio_cell_value, total_qty_cell_value, current_row_index, current_column_index)
        return is_header_row
    
    def get_country_block(self, sheet, row_index, column_index):
        country_block = None
        country_block = sheet.cell(row=row_index, column=column_index-1).value
        return country_block
    
    # def get_size_list(self, sheet, row_index, column_index):
    #     size_list = []
    #     for i in range(column_index+1, column_index+6):
    #         size_list.append(sheet.cell(row=row_index, column=i).value)
    #     return size_list

    def get_total_ratio(self, sheet, row_index, column_index):
        ratio = None
        ratio = sheet.cell(row=row_index+1, column=column_index+6).value
        return ratio
    
    def get_total_qty(self, sheet, row_index, column_index):
        qty = None
        qty = sheet.cell(row=row_index+1, column=column_index+7).value
        return qty
    
    def get_colorway_range_last_row(self, sheet, current_row_index):
        last_row_index = 0
        max_row = sheet.max_row
        for row_index, row in enumerate(sheet.iter_rows(min_row=current_row_index + 1, values_only=True), start=current_row_index + 1):
            if row_index == max_row:
                last_row_index = row_index
                break
            elif row[0] is None:
                last_row_index = row_index - 1
                break
        return last_row_index
    
    def get_colorway_range_last_column(self, sheet, current_row_index):
        max_column_index = 0
        for cell in sheet[current_row_index]:
            if cell.value is not None:
                max_column_index = cell.column
        return max_column_index
    
    def get_style(self, sheet, row_index, column_index):
        style = None
        style = sheet.cell(row=row_index, column=column_index - 1).value
        return style
                        
    def get_colorway(self, sheet, row_index, column_index, ):
        colorway = None
        colorway = sheet.cell(row=row_index, column=column_index).value
        return colorway
    
    def get_sizes(self, sheet, header_row_index, row_index, column_index):
        size_list = []
        for i in range(column_index + 1 , column_index+6):
            size_list.append({'size': sheet.cell(row=header_row_index, column=i).value, 'value': sheet.cell(row=row_index, column=i).value})
        return size_list
    
    def has_none_values(self, sheet, row_index):
        max_column_index = 0
        for cell in sheet[row_index]:
            if cell.value is None:
                continue
            if cell.value is None:
                max_column_index = cell.column
                break
        max_column_index = cell.column
        return max_column_index
    
    def get_processed_po_row_data(self, po_number, costing_version, style_number, country_name, size, ratio, colorway, total_quantity, total_ratio):
        quantity = ratio * (total_quantity/total_ratio)
        size = size.split('/')[1]
        country_object = self.get_matching_object(self._COUNTRY_OBJECT_TYPE, po_number, costing_version=costing_version, country_name=country_name)
        size_object = self.get_matching_object(self._SIZE_OBJECT_TYPE, po_number, costing_version= costing_version, size_name=size)
        data = {
            'country': country_object.country.name if country_object is not None else "N/A",
            'order_country_object': country_object if country_object is not None else None,
            'style': style_number,
            'colorway': colorway,
            'size': size,
            'order_size_object': size_object,
            'quantity': quantity,
        }
        return data
    
    def get_po_data(self, sheet):
        
        grouped_data = {}
        current_po_number = None

        for row_index in range(1, sheet.max_row + 1):
            for column_index in range(1, sheet.max_column + 1):
                self.get_sheet_style(sheet, row_index, column_index)
                self.get_delivery_date(sheet, row_index, column_index)
                if self.is_po_countries_block(sheet, row_index, column_index):
                    po_number = str(self._delivery_date.date()) +'-'+ self._style
                    country_block = self.get_country_block(sheet, row_index, column_index)
                    total_ratio = self.get_total_ratio(sheet, row_index, column_index)
                    total_qty = self.get_total_qty(sheet, row_index, column_index)

                    last_row_in_colorway_range = self.get_colorway_range_last_row(sheet, row_index)
                    last_column_in_colorway_range = self.get_colorway_range_last_column(sheet, row_index)
                    counter = 1
                    for row in sheet.iter_rows(min_row=row_index + 1, max_row=last_row_in_colorway_range, min_col=sheet.min_column, max_col=last_column_in_colorway_range, values_only=True):
                        style = self.get_style(sheet, row_index + counter, column_index)
                        colorway = self.get_colorway(sheet, row_index + counter, column_index)
                        sizes = self.get_sizes(sheet, row_index, row_index + counter, column_index)
                        style_number = self._style
                        if po_number:
                            if po_number not in grouped_data:
                                grouped_data[po_number] = {'costing_version': None, 'delivery_date':None,  'data': []}
                                current_po_number = po_number
                                
                            if not grouped_data[current_po_number]['costing_version']:
                                costing_version = self.get_matching_object(self._COSTING_VERSION_OBJECT_TYPE, current_po_number, style_number=style_number, customer_id=self.customer_id)
                                grouped_data[current_po_number]['costing_version'] = costing_version
                                grouped_data[current_po_number]['delivery_date'] = self._delivery_date
                                current_costing_version = costing_version
            
                            for item in sizes:
                                size = item['size']
                                ratio = item['value']
                                processed_data = self.get_processed_po_row_data(current_po_number, current_costing_version, style_number, country_block, size, ratio, colorway, total_qty, total_ratio)
                                grouped_data[current_po_number]['data'].append(processed_data)
                            counter += 1
                
        return grouped_data
    
    def process_file(self):

        file = handle_file_read(self.excel_file_path)
        workbook = openpyxl.load_workbook(file, data_only=True)
        sheet = workbook.active
        errors = None
        data = self.get_po_data(sheet)
        if not data:
            self.add_error(self.GENERAL_ERRORS_KEY, "Purchase order format is not valid. Unable to locate headers. Please make sure the Item row contains the headers.")
        workbook.close()

        return errors, data

    def process_and_create_purchase_orders(self):
        errors_in_process, data = self.process_file()
        errors = []
        if errors_in_process:
            errors.append(errors_in_process)

        po_ids = self.create_purchase_orders(data, self.customer_id, self.attachment_id)
        response = {
            'created': not self.have_errors,
        }
        if not self.have_errors:
            response['po_list'] = po_ids
        else:
            response['errors'] = self.errors
        return response
