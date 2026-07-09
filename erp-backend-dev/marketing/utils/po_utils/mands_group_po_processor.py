import openpyxl
from datetime import datetime
from marketing.utils.aws_utils import handle_file_read
from marketing.mixins.po_processor_mixins import POProcessorMixin


class MandSGroupPOProcessor(POProcessorMixin):
    _fob_text = ['fob', 'Fob', 'FOB']
    _channel_text = ['channel', 'Channel']
    _avail_text = ['avail', 'Avail']
    _quantity_text = ['quantity', 'Quantity']
    po_counter = 1

    def __init__(self, excel_file_path, customer_id, attachment_id):
        super().__init__()
        self.excel_file_path = excel_file_path
        self.customer_id = customer_id
        self.attachment_id = attachment_id

    def santize_value(self, value):
        return str(value).lower().strip()

    def is_po_colorways_block(self, sheet, current_row_index, current_column_index):
        has_errors = False
        is_colorways_row = False
        fob_row_cell_value = self.santize_value(sheet.cell(row=current_row_index, column=current_column_index + 1).value)
        channel_cell_value = self.santize_value(sheet.cell(row=current_row_index, column=current_column_index + 2).value)
        quantity_cell_value = self.santize_value(sheet.cell(row=current_row_index, column=current_column_index + 3).value)
        if fob_row_cell_value in self._fob_text and channel_cell_value in self._channel_text and \
            quantity_cell_value in self._quantity_text:
            # print(fob_row_cell_value,channel_cell_value,quantity_cell_value)
            is_colorways_row = True
        return is_colorways_row
    
    def get_header_idx(self, sheet):
        headers = [cell.value for cell in sheet[1]]
        header_indices = {header: idx for idx, header in enumerate(headers)}
        has_errors = False

        fob_column_idx = None
        channel_column_idx = None
        avail_column_idx = None
        quantity_column_idx = None
        
        for fob_header in self._fob_text:
            if fob_header in header_indices:
                fob_column_idx = header_indices[fob_header]
                break
        for channel_header in self._channel_text:
            if channel_header in header_indices:
                channel_column_idx = header_indices[channel_header]
                break
        for avail_header in self._avail_text:
            if avail_header in header_indices:
                avail_column_idx = header_indices[avail_header]
                break
        for quantity_header in self._quantity_text:
            if quantity_header in header_indices:
                quantity_column_idx = header_indices[quantity_header]
                break
        if fob_column_idx is None or channel_column_idx is None or avail_column_idx is None or quantity_column_idx is None:
            has_errors = True
        else:
            has_errors = False
        return has_errors, fob_column_idx, channel_column_idx, avail_column_idx, quantity_column_idx
    
    def get_colorway_range_last_row(self, sheet, current_row_index):
        last_row_index = 0
        max_row = sheet.max_row
        for row_index, row in enumerate(sheet.iter_rows(min_row=current_row_index + 1, values_only=True), start=current_row_index + 1):
            if row_index == max_row:
                last_row_index = row_index
                break
            elif row[0] is not None:
                last_row_index = row_index -1
                break
        return last_row_index
    
    def get_colorway_range_last_column(self, sheet, current_row_index):
        max_column_index = 0
        for cell in sheet[current_row_index]:
            if cell.value is not None:
                max_column_index = cell.column
        return max_column_index
    
    def has_none_values(self, sheet, row_index):
        max_column_index = 0
        for cell in sheet[row_index]:
            if cell.value is None:
                continue
            if cell.value is None:
                max_column_index = cell.column
                break
        max_column_index = cell.column
        return max_column_index
    
    def get_colorway(self, sheet, po_details_start_row_index, column_index):
        colorway = None
        new = column_index-2

        # print(new)
        colorway = sheet.cell(row=po_details_start_row_index, column=new).value
        return colorway
    
    def get_po_data(self, sheet):
        has_error, fob_column_idx, channel_column_idx, avail_column_idx, quantity_column_idx = self.get_header_idx(sheet)
        
        grouped_data = {}
        current_style_number = None
        current_po_number = None
        delivery_date = None

        for row_index in range(1, sheet.max_row + 1):
            for column_index in range(1, sheet.max_column + 1):
                if self.is_po_colorways_block(sheet, row_index, column_index):
                    # print(row_index, column_index)
                    current_style_number = None
                    current_po_number = None
                    current_costing_version = None
                    previous_po_number = None

                    last_row_in_colorway_range = self.get_colorway_range_last_row(sheet, row_index)
                    last_column_in_colorway_range = self.get_colorway_range_last_column(sheet, row_index)
                    colorway = self.get_colorway(sheet, row_index, column_index)
                    
                    for row in sheet.iter_rows(min_row=row_index, max_row=last_row_in_colorway_range, min_col=sheet.min_column, max_col=last_column_in_colorway_range, values_only=True):
                        if row[fob_column_idx] is not None and row[channel_column_idx] is not None and row[quantity_column_idx] is not None:
                            if row[0] is None:
                                quantities = row[quantity_column_idx+1:last_column_in_colorway_range + 1]
                                country = None
                                delivery_date = row[fob_column_idx]
                                current_datetime = datetime.now()
                                po_number = f"PO-{current_datetime.strftime('%Y%m%d%H%M%S')}-{self.po_counter}"

                                for column_idx in range(len(quantities)):
                                    size = sheet.cell(row=row_index, column=column_idx+1 + quantity_column_idx+1).value
                                    quantity = quantities[column_idx]
                                    # print(quantities)
                                    if size is not None:
                                        # print("Size")
                                        if po_number:
                                            if po_number not in grouped_data:
                                                grouped_data[po_number] = {'costing_version': None, 'delivery_date':None,  'data': []}
                                                # print("sadsadas")
                                                current_po_number = po_number

                                        if not grouped_data[current_po_number]['costing_version']:
                                            # print(current_po_number)
                                            costing_version = self.get_matching_object(self._COSTING_VERSION_OBJECT_TYPE, current_po_number, style_number=colorway, customer_id=self.customer_id)
                                            grouped_data[current_po_number]['costing_version'] = costing_version
                                            grouped_data[current_po_number]['delivery_date'] = delivery_date
                                            current_style_number = colorway
                                            current_costing_version = costing_version
                                            if not costing_version:
                                                break
                                            previous_po_number = po_number
                                        elif current_style_number != colorway:
                                            self.add_error(current_po_number, "PO Number %s has multiple Style Numbers. A PO can only have one Style Number")

                                        processed_data = self.get_processed_po_row_data(current_po_number, current_costing_version, colorway, country, size, colorway, quantity)
                                        grouped_data[current_po_number]['data'].append(processed_data)
                                        # print(grouped_data)
                                        self.po_counter += 1
        # print(grouped_data)
        return grouped_data
    

    def process_file(self):

        file = handle_file_read(self.excel_file_path)
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        data = {}
        errors = None
        data = {}
        data = self.get_po_data(sheet)
        workbook.close()

        return errors, data

    def process_and_create_purchase_orders(self):
        errors_in_process, data = self.process_file()
        errors = []
        if errors_in_process:
            errors.append(errors_in_process)

        po_ids = self.create_purchase_orders(data, self.customer_id, self.attachment_id)
        response = {
            'created': not self.have_errors,
        }
        if not self.have_errors:
            response['po_list'] = po_ids
        else:
            response['errors'] = self.errors
        return response
