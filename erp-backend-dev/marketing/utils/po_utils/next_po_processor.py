import openpyxl
from marketing.utils.aws_utils import handle_file_read
from marketing.models import PurchaseOrder
from marketing.mixins.po_processor_mixins import POProcessorMixin
import datetime


class NextPOProcessor(POProcessorMixin):
    _contract_number_text = ['contract no', 'contract no.', 'contract number']
    _delivery_number_text = ['del']
    _item_option_text = ['item/option', ]
    _size_text = 'size'
    _quantity_text = 'quantity'
    _style_text = ['Item', 'item']
    _delivery_date_text = ['Ex-Fact', 'ex-fact']
    _colorway_text = ['Description', 'description']

    def purchase_order_exists(self, po_number, customer_id):
        order_exists = PurchaseOrder.objects.filter(name=po_number, customer_id=self.customer_id)
        return order_exists.exists()

    def __init__(self, excel_file_path, customer_id, attachment_id):
        super().__init__()
        self.excel_file_path = excel_file_path
        self.customer_id = customer_id
        self.attachment_id = attachment_id

    def santize_value(self, value):
        return str(value).lower().strip()
    
    def formatted_date(self, incomplete_date):
        current_year = datetime.datetime.now().year
        month, day = map(int, incomplete_date.split('/'))
        new_date = datetime.date(current_year, month, day)
        return new_date

    def is_po_style_block(self, sheet, current_row_index, current_column_index):
        style_number = None
        current_cell_value = sheet.cell(row=current_row_index, column=current_column_index).value
        if current_cell_value in self._style_text:
            for index in range(1, 5):
                next_row_index = current_row_index + index
                style_number = sheet.cell(row=next_row_index, column=current_column_index).value
                if style_number is not None:
                    return style_number
        else:
            (self.GENERAL_ERRORS_KEY, "Purchase order format is not valid. Unable to locate headers. Please make sure after the item cell contains the style.")

    # CHange this function
    def is_po_colorway_block(self, sheet, current_row_index, current_column_index):
        colorway = None
        current_cell_value = sheet.cell(row=current_row_index, column=current_column_index).value
        if current_cell_value in self._colorway_text:
            for index in range(1, 5):
                next_row_index = current_row_index + index
                colorway = sheet.cell(row=next_row_index, column=current_column_index).value
                if colorway is not None:
                    return colorway
        else:
            (self.GENERAL_ERRORS_KEY, "Purchase order format is not valid. Unable to locate headers. Please make sure after the item cell contains the style.")

    def is_po_quantities_block(self, sheet, current_row_index, current_column_index):
        size_row_cell_value = self.santize_value(sheet.cell(row=current_row_index + 1, column=current_column_index).value)
        quanity_row_cell_value = self.santize_value(sheet.cell(row=current_row_index + 2, column=current_column_index).value)
        current_cell_value = self.santize_value(sheet.cell(row=current_row_index, column=current_column_index).value)
        is_quantities_row = False
        if current_cell_value in self._item_option_text and size_row_cell_value == self._size_text and quanity_row_cell_value == self._quantity_text:
            is_quantities_row = True
        return is_quantities_row

    def get_contract_number(self, sheet, po_details_start_row_index, po_details_start_column_index):
        process_rows = po_details_start_row_index + 3
        contract_number = None
        for row_index in range(po_details_start_row_index, process_rows):
            for cell_index in range(1, po_details_start_column_index + 1):
                if self.santize_value(sheet.cell(row=row_index, column=cell_index).value) in self._contract_number_text:
                    contract_number = sheet.cell(row=row_index + 1, column=cell_index).value
                    break
                cell_index += 1
        return contract_number
    
    def get_style_number(self, sheet, po_details_start_row_index, po_details_start_column_index):
        process_rows = po_details_start_row_index + 3
        style_number = None
        for row_index in range(po_details_start_row_index, process_rows):
            for cell_index in range(1, po_details_start_column_index + 1):
                if self.santize_value(sheet.cell(row=row_index, column=cell_index).value) in self._style_text:
                    style_number = sheet.cell(row=row_index + 4, column=cell_index).value
                    break
                cell_index += 1
        return style_number
    
    def get_delivery_number(self, sheet, po_details_start_row_index, po_details_start_column_index):
        process_rows = po_details_start_row_index + 3
        delivery_number = None
        for row_index in range(po_details_start_row_index, process_rows):
            for cell_index in range(1, po_details_start_column_index + 1):
                if self.santize_value(sheet.cell(row=row_index, column=cell_index).value) in self._delivery_number_text:
                    delivery_number = sheet.cell(row=row_index + 1, column=cell_index).value
                    break
                cell_index += 1
        return delivery_number
    
    def get_delivery_date(self, sheet, po_details_start_row_index, po_details_start_column_index):
        process_rows = po_details_start_row_index + 3
        _delivery_date_text = None
        for row_index in range(po_details_start_row_index, process_rows):
            for cell_index in range(1, po_details_start_column_index + 1):
                if self.santize_value(sheet.cell(row=row_index, column=cell_index).value) in self._delivery_date_text:
                    _delivery_date_text = sheet.cell(row=row_index + 1, column=cell_index).value
                    break
                cell_index += 1
        return _delivery_date_text

    def get_po_data(self, sheet):
        have_errors = False
        grouped_data = {}
        current_style_number = None
        current_po_number = None
        current_costing_version = None
        previous_contract_number = None
        current_delivery_date = None

        for row_index in range(1, sheet.max_row + 1):
            for column_index in range(1, sheet.max_column + 1):
                if self.is_po_style_block(sheet, row_index, column_index):
                    style_number = self.is_po_style_block(sheet, row_index, column_index)
                if self.is_po_colorway_block(sheet, row_index, column_index):
                    colorway = self.is_po_colorway_block(sheet, row_index, column_index)
                if self.is_po_quantities_block(sheet, row_index, column_index):
                    if not style_number:
                        have_errors = True
                    contract_number = self.get_contract_number(sheet, row_index, column_index)
                    delivery_number = self.get_delivery_number(sheet, row_index, column_index)
                    delivery_date = self.get_delivery_date(sheet, row_index, column_index)

                    if delivery_date is not None:
                        current_delivery_date = self.formatted_date(delivery_date)

                    if delivery_number is not None:
                        contract_number = contract_number + '-' + str(delivery_number)
                    if not contract_number:
                        contract_number = previous_contract_number

                    for column_index_in_po_block in range(column_index + 1, sheet.max_column + 1):
                        po_number = contract_number
                        if po_number:
                            if po_number not in grouped_data:
                                grouped_data[po_number] = {'costing_version': None, 'delivery_date':None,  'data': []}
                                current_po_number = po_number

                        if colorway:
                            colorway = colorway
                            size = sheet.cell(row=row_index + 1, column=column_index_in_po_block).value
                            quantity = sheet.cell(row=row_index + 2, column=column_index_in_po_block).value
                            country = None

                            if not grouped_data[current_po_number]['costing_version']:
                                costing_version = self.get_matching_object(self._COSTING_VERSION_OBJECT_TYPE, current_po_number, style_number=style_number, customer_id=self.customer_id)
                                grouped_data[current_po_number]['costing_version'] = costing_version
                                grouped_data[current_po_number]['delivery_date'] = current_delivery_date
                                current_style_number = style_number
                                current_costing_version = costing_version
                                if not costing_version:
                                    break
                            elif current_style_number != style_number:
                                self.add_error(current_po_number, "PO Number %s has multiple Style Numbers. A PO can only have one Style Number")
                            processed_data = self.get_processed_po_row_data(current_po_number, current_costing_version, style_number, country, size, colorway, quantity)
                            grouped_data[current_po_number]['data'].append(processed_data)
                        else:
                            break
                    previous_contract_number = contract_number
        return grouped_data

    def process_file(self):
        file = handle_file_read(self.excel_file_path)
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        errors = None
        data = {}
        
        data = self.get_po_data(sheet)
        if not data:
            self.add_error(self.GENERAL_ERRORS_KEY, "Purchase order format is not valid. Unable to locate headers. Please make sure the Item row contains the headers.")

        workbook.close()
        return errors, data

    def process_and_create_purchase_orders(self):
        errors_in_process, data = self.process_file()
        errors = []
        if errors_in_process:
            errors.append(errors_in_process)

        po_ids = self.create_purchase_orders(data, self.customer_id, self.attachment_id)
        response = {
            'created': not self.have_errors,
        }
        if not self.have_errors:
            response['po_list'] = po_ids
        else:
            response['errors'] = self.errors
        return response