import os
import re
from abc import abstractmethod
import secrets
from tempfile import NamedTemporaryFile

from django.core.files import File
from django.core.files.storage import FileSystemStorage
from django.core.files.uploadedfile import InMemoryUploadedFile

from materials.fieldmetadata import material_metadata
from bs4 import BeautifulSoup

from materials.fieldmetadata.material_metadata import ATTRIBUTE_DISPLAY_VALUE_KEY, get_embellishment_headers, \
    get_wash_headers
from materials.fieldmetadata.measuring_unit_helpers import MaterialUnitHelper, PerMeasuringUnitHelper
from shared import email
import uuid
from marketing.models import OrderInquiry, OrderCostingVersion, PackItemService, OrderInquiryProgram
from materials.models import SupplierInquiry, SupplierInquiryEmail, UserDefinedMaterial, GenericMaterial, Material, \
    SupplierInquiryDetail, SupplierInquiryMaterialCode
from django.template.loader import render_to_string

from shared.models import Supplier, CREDIT_DAYS_30_PAYMENT_METHOD_TYPE
from shared.utils import get_object_or_none
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from marketing.utils.aws_utils import handle_uploaded_file, handle_file_read
from io import BytesIO
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from shared.helpers.converters import choice_to_dictionary, choice_to_reverse_dictionary


def read_email_body(body):
    manual_data_enter = False
    # ************** function for get table data *******************
    def get_table_data(table):
        rows = table.find_all('tr')
        row_count = len(rows)
        table_headers = []
        table_rows = []
        if row_count > 1:
            for header in rows[0].find_all('th'):
                header_name = header.text.lower().replace(' ', '_')
                if header_name == 'cif':
                    header_name = 'cif_price'
                if '/' in header_name:
                    header_name = header_name.split('_')[1]
                table_headers.append(
                    {'label': header.text, 'name': header_name})
            for data_row_index in range(1, row_count):

                row_data = rows[data_row_index].find_all('td')
                table_rows.append(
                        {table_headers[data_column_index]['name']: row_data[data_column_index].text
                         for data_column_index in range(0, len(row_data))})
            return table_rows
        else:
            # TODO Throw a error flag to manual data entering
            return table_rows
    #******************** function for get date *********************
    def get_date(text, lookup_string):
        date = ""
        if lookup_string in text:
            date_string = text.split(':-')[1].split('/')
            if len(date_string) == 3:
                date_error = False
                for value in date_string:
                    if not value.isnumeric():
                        date_error = True
                        return "error"
                if not date_error:
                    DD = date_string[0]
                    MM = date_string[1]
                    YY = date_string[2]
                    if int(DD)>0 and int(DD)<=31 and int(MM)>0 and int(MM)<=12:
                        date = YY + '-' + MM + '-' + DD
                        return date
                    else:
                        return "error"
            else:
                return ""

    
    ################ MAIN CODE #######################
    soup = BeautifulSoup(body)
    date_lookups = {
        'expiration_date': 'Expiration Date',
        'tentative_material_in_house_date': 'Tentative Material In House Date'
    }
    dates={}

    for p in soup.find_all('p'):
        if 'MSG_CODE_' in p.text:
            msg_code = p.text.replace('MSG_CODE_','')
            if len(msg_code) == 12:
                pass # update email that sent related to this email
        for key in date_lookups:
            date = get_date(text=p.text, lookup_string=date_lookups[key])
            if date == 'error':
                manual_data_enter =True
            elif not date == None:
                dates[key] = date
    if not len(dates) == 2:
        manual_data_enter = True
    for table in soup.find_all('table'):
        
        for table_row in get_table_data(table):
            if not table_row == []:
                try:
                    supplier_inquiry = SupplierInquiry.objects.get(hash_id = table_row['hash_id'])
                    attribute_list = ['fob_price',
                                    'freight_charge',
                                    'ship_mode',
                                    'cif_price',
                                    'cutting_width',
                                    'cutting_width_unit',
                                    'costing_unit']
                    error = False
                    updated = False
                    supplier_inquiry.has_supplier_feedback = True
                    for key in attribute_list:
                        if key in table_row:
                            updated = True
                            value = table_row[key]
                            if not value == "":
                                setattr(supplier_inquiry, key, value)
                            else:
                                error = True
                    if updated:
                        if error:
                            supplier_inquiry.email_status = SupplierInquiry.RECEIVED_NEED_REVIEW
                        else:
                            supplier_inquiry.email_status = SupplierInquiry.RECEIVED_AND_PROCESSED
                    else:
                        supplier_inquiry.email_status = SupplierInquiry.RECEIVED_AND_PROCESSED
                    for key in dates:
                        setattr(supplier_inquiry, key, dates[key])
                    supplier_inquiry.email_status = SupplierInquiry.PENDING_EMAIL
                    supplier_inquiry.save()

                except SupplierInquiry.DoesNotExist:
                    print('Supplier Inquiry Cannot Find')
            else:
                manual_data_enter = True


class InquiryEmailSender:
    highlighted_columns = ['fob_price', 'cif_price', 'ex_work_price', 'ship_mode', 'cutting_width', 'lead_time',
                           'cutting_width_unit', 'pay_mode', 'costing_unit', ]
    quote_fields = ['lead_time', 'fob_price', 'cif_price', 'ship_mode', 'ex_work_price', 'pay_mode'] # TODO - change this to use the proper fields

    @staticmethod
    def get_supplier_inquiry_headers():
        headers = []
        for h in material_metadata.get_supplier_inquiry_quote_meta_data():
            if h[ATTRIBUTE_DISPLAY_VALUE_KEY] in InquiryEmailSender.quote_fields:
                headers.append(h)
        return headers

    @staticmethod
    def get_service_headers(service_type):
        main_headers = [
            {'label': 'Hash ID', 'name': 'hash_id'},
        ]
        service_headers = []
        if service_type == PackItemService.EMBELLISHMENT_SERVICE_TYPE:
            service_headers = get_embellishment_headers()
        elif service_type == PackItemService.WASH_SERVICE_TYPE:
            service_headers = get_wash_headers()

        for header in service_headers:
            main_headers.append(header)

        inquiry_headers = InquiryEmailSender.get_supplier_inquiry_headers()
        main_headers = [*main_headers, *inquiry_headers]
        return main_headers

    @staticmethod
    def get_material_headers(material_type):
        main_headers = [
            {'label': 'Hash ID', 'name': 'hash_id'},
            {'label': 'Customer Reference Code', 'name': 'customer_reference_code'},
            {'label': 'RITZ Code', 'name': 'ritz_code'},
            {'label': 'Supplier Reference Code', 'name': 'supplier_reference_code'},
            {'label': 'Estimated Quantity', 'name': 'estimated_quantity_display'},
            {'label': 'MOQ', 'name': 'minimum_order_quantity'},
            {'label': 'MOQ Unit', 'name': 'minimum_order_quantity_units'}
            # {'label': 'Customer', 'name': 'customer'},
            # {'label': 'Brand', 'name': 'brand'},
        ]
        material_headers = UserDefinedMaterial.get_material_headers(material_type)
        for material_header in material_headers:
            main_headers.append(material_header)

        if material_type == 'fabric':
            # main_headers.append({'label': 'Costing Unit', 'name': 'costing_unit', 'width': "60px"})
            main_headers.append({'label': 'Cutting Width', 'name': 'cutting_width', 'width': "60px"})
            main_headers.append({'label': 'Cutting Width Unit', 'name': 'cutting_width_unit', 'width': "60px"})
        main_headers.append({'label': 'Costing Unit', 'name': 'costing_unit', 'width': "60px"})
        # Add supplier quote headers
        supplier_inquiry_headers = InquiryEmailSender.get_supplier_inquiry_headers()
        main_headers = [*main_headers, *supplier_inquiry_headers]
        return main_headers

    def flatten_table_data(self, material_data):
        for material_type, material_info in material_data.items():
            table_data = material_info['table_data']
            if material_info['is_material']:
                data_list = []

                for customer_brand_material_id, material_attriibutes in table_data.items():
                    data_list.append(material_attriibutes)
                material_data[material_type]['table_data'] = data_list
        return material_data
    
    def create_excel_file(self,template_data):
        # print(template_data['data_array']['carelabel']['table_headers'])
        from django.conf import settings
        hide_cols = template_data['hide_columns']
        base_dir = settings.BASE_DIR
        file_path = "data_files/excel/price_inquiry_for_ritz_clothing_template.xlsx"
        excel_template = os.path.join(base_dir, file_path)

        output={}
        excel_file = openpyxl.load_workbook(
            filename=excel_template,
            read_only=False
        )
        sheet = excel_file.active
        sheet.title = 'Price Inquiry'
        ritz_header_col = 8
        excel_row = 8
        excel_col = 1
        thin_border = openpyxl.styles.Side(border_style='thin', color='000000')
        border = openpyxl.styles.Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
        # print(template_data)
        validation_list, choice_list = SupplierInquiry().get_lists_of_validation_for_excel()
        default_values = {'ship_mode': SupplierInquiryDetail.SEA_CHOICE, 'pay_mode': CREDIT_DAYS_30_PAYMENT_METHOD_TYPE, 'cutting_width_unit': MaterialUnitHelper.INCHES_UNIT, 'costing_unit': PerMeasuringUnitHelper.PER_INCH_UNIT, 'minimum_order_quantity_units': MaterialUnitHelper.METERS_UNIT}
        validation_list_default_values = {key: choice_to_dictionary(choice_list[key])[default_values[key]] for key in default_values}
        for material in template_data['data_array']:
            material_data = template_data['data_array'][material]

            excel_col=1
            sheet.cell(row=excel_row, column=excel_col).value=material_data['material_type']
            sheet.column_dimensions[get_column_letter(excel_col)].width = 25
            sheet.cell(row=excel_row, column=excel_col).alignment = Alignment(horizontal='right')
            excel_row+=1
            excel_col+=1
            table_data_start_row = excel_row
            table_data_start_col = excel_col
            table_data_end_row = excel_row
            table_data_end_col = excel_col
            for table_header in material_data['table_headers']:
                excel_row = table_data_start_row
                # sheet.column_dimensions[get_column_letter(excel_col)].auto_size = True
                sheet.cell(row=excel_row, column=excel_col).value = table_header['label']
                sheet.column_dimensions[get_column_letter(excel_col)].width = len(table_header['label'])*1.5
                # sheet.cell(row=excel_row, column=excel_col).auto_size = True
                excel_row+=1
                table_header_key = table_header.get('name', None)
                if table_header_key in validation_list_default_values:
                    data_val = DataValidation(type="list", formula1=validation_list[table_header_key])
                    sheet.add_data_validation(data_val)
                    for temp_row in range(excel_row,table_data_end_row+1):
                        data_val.add(sheet.cell(row=temp_row, column=excel_col))
                        sheet.cell(row=temp_row, column=excel_col).value = validation_list_default_values[table_header_key]
                else:
                    for table_data in material_data['table_data']:
                        if table_header_key in table_data:
                            sheet.cell(row=excel_row, column=excel_col).value = table_data[table_header_key]
                            if table_header_key in hide_cols:
                                sheet.column_dimensions[get_column_letter(excel_col)].hidden= True
                        elif table_header.get('attachment_field_name', None) in table_data:
                            # print(table_data[table_header['attachment_field_name']])
                            # print(table_data)
                            if table_data['embellishment_attachment'] != {}:
                                try:
                                    img = Image(handle_file_read(table_data['embellishment_attachment']['file_path']))
                                    img.width = img.width/4
                                    img.height = img.height/4
                                    rd = sheet.row_dimensions[excel_row]
                                    rd.height = img.height/1.1
                                    rd.alignment = Alignment(horizontal='center')
                                    cd = sheet.column_dimensions[get_column_letter(excel_col)]
                                    cd.width = img.width/5
                                    sheet.add_image(img,get_column_letter(excel_col)+str(excel_row))
                                    sheet.cell(row=excel_row, column=excel_col).hyperlink = table_data['embellishment_attachment']['file_path']
                                    sheet.cell(row=excel_row, column=excel_col).style = "Hyperlink"
                                except:
                                    print(table_data['embellishment_attachment']['file_path'])
                            sheet.cell(row=excel_row, column=excel_col).value = table_data['sub_type'] # TODO - make sure it is linked

                        if table_data_end_row < excel_row:
                            table_data_end_row = excel_row
                        excel_row+=1
                if table_data_end_col < excel_col:
                    table_data_end_col = excel_col
                cd = sheet.column_dimensions[get_column_letter(excel_col)]
                # cd.auto_size = True
                excel_col+=1
                if ritz_header_col < excel_col+1:
                    sheet.insert_cols(ritz_header_col)
                    ritz_header_col+=1
            # print(get_column_letter(table_data_start_col)+str(table_data_start_row)+':'+get_column_letter(table_data_end_col)+str(table_data_end_row))
            # table_range = sheet[get_column_letter(table_data_start_col)+str(table_data_start_row)+':'+get_column_letter(table_data_end_col)+str(table_data_end_row)]
            # print(table_data_end_col)
            for col in range(table_data_start_col,table_data_end_col+1):
                for row in range(table_data_start_row,table_data_end_row+1):
                    sheet.cell(row=row, column=col).border = border
            # table_range.border = border
            excel_row+=1
        sheet.delete_cols(ritz_header_col-1)
        # excel_file.save('Price Inquiry For RITZ Clothing.xlsx')

        output['mime_type'] = excel_file.mime_type
        # print("excel created")
        with NamedTemporaryFile(mode='w+', encoding='utf8', delete=False, suffix='.xlsx') as tmp:
            # print(tmp.name)
            excel_file.save(tmp.name)
            # print("test 1.2")
            file = open(tmp.name, mode='rb')
            # file_obj = File(file=file)
            # print(self.get_file_save_path())
            saved_file = handle_uploaded_file(file, self.get_file_save_path(), 'Price Inquiry For RITZ Clothing.xlsx')
            output['path'] = saved_file
            # output['path'] = file.name
            file.close()
            # file_obj.close()
        excel_file.close()
        # print("excel saved")
        return output
        
    def process_emails(self, suppliers):
        cc_emails = self.get_cc_emails()
        # print(suppliers)
        for supplier in suppliers:
            # print(supplier.name)
            style_numbers = []
            data_set = {}
            email_pending_supplier_inquiries = self.get_pending_supplier_inquiries(supplier)
            if not email_pending_supplier_inquiries.exists():
                continue
            processed_supplier_inquiry_ids = []
            for supplier_inquiry in email_pending_supplier_inquiries:
                # print(supplier_inquiry.id)
                processed_supplier_inquiry_ids.append(supplier_inquiry.id)
                

                if supplier_inquiry.version:
                    if supplier_inquiry.version.order.style_number not in style_numbers:
                        style_numbers.append(supplier_inquiry.version.order.style_number)
                
                if supplier_inquiry.customer_brand_material and supplier_inquiry.customer_brand_material.material_type:
                    type = supplier_inquiry.customer_brand_material.material_type
                    if not type in data_set:
                        data_set[type] = {}
                        data_set[type]['is_material'] = True
                        data_set[type]['table_headers'] = self.get_material_headers(type)
                        data_set[type]['table_data'] = {}

                        # Add material headers
                        attributes = supplier_inquiry.customer_brand_material.get_attributes()
                        data_set[type]['material_type'] = attributes['material_label']
                        data_set[type]['material_name'] = attributes[GenericMaterial.MATERIAL_TYPE_KEY]
                    # supplier_iquiry_material_code = get_object_or_none(SupplierInquiryMaterialCode, {'customer_brand_material': supplier_inquiry.customer_brand_material,
                    #                                                                                  'supplier': supplier_inquiry.supplier})
                    supplier_iquiry_material_code = SupplierInquiryMaterialCode.objects.filter(
                        customer_brand_material=supplier_inquiry.customer_brand_material,
                        supplier=supplier_inquiry.supplier
                    )
                    if supplier_iquiry_material_code.exists():
                        supplier_iquiry_material_code=supplier_iquiry_material_code[0]
                    
                    if supplier_iquiry_material_code:
                        supplier_iquiry_reference_code = supplier_iquiry_material_code.supplier_material_reference_code
                        if not supplier_iquiry_reference_code:
                            supplier_iquiry_reference_code = ''
                    else:
                        supplier_iquiry_reference_code = ''
                    attributes = supplier_inquiry.customer_brand_material.get_attributes()
                    material_data = {
                        'hash_id': supplier_inquiry.hash_id,
                        'customer_reference_code': supplier_inquiry.customer_brand_material.material_code.customer_reference_code if supplier_inquiry.customer_brand_material.material_code else '',
                        'ritz_code': supplier_inquiry.customer_brand_material.verbose_reference_code,
                        'supplier_reference_code': supplier_iquiry_reference_code,
                        'minimum_order_quantity': '',
                        'minimum_order_quantity_units': '',
                        **attributes
                    }

                    current_data = data_set[type]['table_data'].get(supplier_inquiry.customer_brand_material_id, None)
                    estimated_quantity = supplier_inquiry.customer_brand_material.get_version_estimated_quantity(supplier_inquiry.version)

                    # If estimated quantity is None for current or a previous one, it is set to None
                    if current_data and current_data.get('estimated_quantity'):
                        if estimated_quantity['estimated_quantity']:
                            combined_quantity = current_data['estimated_quantity'] + estimated_quantity['estimated_quantity']
                            data_set[type]['table_data'][supplier_inquiry.customer_brand_material_id]['estimated_quantity'] = combined_quantity
                            data_set[type]['table_data'][supplier_inquiry.customer_brand_material_id]['estimated_quantity_display'] = '%s %s' % (combined_quantity, estimated_quantity['estimated_quantity_units'])

                        else:
                            data_set[type]['table_data'][supplier_inquiry.customer_brand_material_id]['estimated_quantity'] = None
                            data_set[type]['table_data'][supplier_inquiry.customer_brand_material_id]['estimated_quantity_display'] = "N/A"
                    else:
                        # If current data by no estimated quantity in previously added data
                        if current_data:
                            current_data['hash_id'] = "%s,%s" % (str(current_data['hash_id']), supplier_inquiry.hash_id)
                            material_data = {
                                **current_data,
                                **estimated_quantity,
                                'estimated_quantity_display': estimated_quantity.get('estimated_quantity_display', 'N/A')
                            }
                        else:
                            material_data = {
                                **material_data,
                                **estimated_quantity,
                                'estimated_quantity_display': estimated_quantity.get('estimated_quantity_display', 'N/A')
                            }

                    data_set[type]['table_data'][supplier_inquiry.customer_brand_material_id] = material_data
                elif supplier_inquiry.item_service:
                    type = supplier_inquiry.item_service.service_type
                    if not data_set.get(type, None):
                        data_set[type] = {}
                        data_set[type]['is_material'] = False
                        data_set[type]['table_headers'] = self.get_service_headers(type)
                        data_set[type]['table_data'] = []
                        data_set[type]['material_type'] = supplier_inquiry.item_service.get_service_type_display()
                    attributes = supplier_inquiry.item_service.get_attributes()
                    material_data = {
                        'hash_id': supplier_inquiry.hash_id,
                        'minimum_order_quantity': '',
                        'minimum_order_quantity_units': '',
                        **attributes
                    }
                    data_set[type]['table_data'].append(material_data)
                    processed_supplier_inquiry_ids.append(supplier_inquiry.id)
                supplier_inquiry.save()

            hide_columns = ['hash_id']

            data_set = self.flatten_table_data(data_set)
            template_data = {
                'data_array': data_set,
                'highlighted_columns': self.highlighted_columns,
                'hide_columns': hide_columns,
                'fabric_loop': [1, 2],
                **self.get_customer_brand()
            }
            body = render_to_string('../templates/supplier_inquiry_email_templet.html', template_data)
            # print(template_data)
            style_nums = ", ".join(style_numbers)
            subject = 'Price Inquiry For RITZ Clothing (Style Number(s): %s)' % style_nums
            # print("Test 1")
            # print(template_data)
            excel_file=self.create_excel_file(template_data)
            # print("Test 2")
            sent_email = email.send_email(
                to=[supplier.email], cc=cc_emails, subject=subject,
                 body=body.replace('\n', ''), attachments=[excel_file])
            # TODO - is this needed ?
            supplier_inquiry_email = SupplierInquiryEmail.objects.create(email=sent_email)
            for supplier_inquiry in email_pending_supplier_inquiries:
                supplier_inquiry_email.supplier_inquiry.add(supplier_inquiry)
                supplier_inquiry.email_status = supplier_inquiry.PENDING_RESPONSE

            SupplierInquiry.objects.filter(pk__in=processed_supplier_inquiry_ids).update(email_status=SupplierInquiry.PENDING_RESPONSE) # comment this out when debugging to make it easier
            supplier_inquiry_email.save() # Uncomment for saving sending status

            # file_html = open("supplier_inquiry_email_demo.html", "w")
            # file_html.write(body.replace('\n', ''))
            # file_html.close()

    @abstractmethod
    def get_pending_supplier_inquiries(self, supplier):
        ...

    @abstractmethod
    def get_customer_brand(self):
        ...

    @abstractmethod
    def get_cc_emails(self):
        ...

    @abstractmethod
    def get_file_save_path(self):
        ...


class OrderVersionSupplierInquiryEmailSender(InquiryEmailSender):

    def __init__(self, version_id):
        self.version_id = version_id
        self.version = OrderCostingVersion.objects.get(pk=version_id)

    def get_file_save_path(self):
        unique_path = secrets.token_hex(15) # see notes
        path = self.version.get_file_saving_path() + 'supplier_inquiries/emails/' + unique_path
        return path

    def get_pending_suppliers(self):
        supplier_ids = SupplierInquiry.objects.filter(version=self.version_id).values_list('supplier_id', flat=True).distinct('supplier')
        suppliers = Supplier.objects.filter(pk__in=supplier_ids)  # get unique suppliers of order_id
        return suppliers

    def get_pending_supplier_inquiries(self, supplier):
        email_pending_supplier_inquiries = SupplierInquiry.objects.filter(
            version=self.version_id,
            supplier=supplier, email_status=SupplierInquiry.QUEUED_EMAIL,
            # version__supplier_inquiries_complete=False
        )
        return email_pending_supplier_inquiries

    def get_cc_emails(self):
        cc = [user.email for user in self.version.order.watchlist.all()]
        if self.version.order.merchant:
            cc.append(self.version.order.merchant.email)
        return cc

    def get_customer_brand(self):
        data = {
            'customer': self.version.order.customer.name,
            'brand': self.version.order.brand.name,
        }
        return data

    def send_supplier_inquiry_email(self):

        # Send only if it is not part of a program
        if not self.version.order.order_program:
            suppliers = self.get_pending_suppliers()

            if not suppliers.exists():
                return {'status': 'No Iquiries to Send Email'}
            self.process_emails(suppliers)


class OrderProgramSupplierInquiryEmailSender(InquiryEmailSender):

    def __init__(self, program_id):
        self.program_id = program_id
        self.program = OrderInquiryProgram.objects.get(pk=program_id)
        self.program_version_ids = self.get_program_version_ids()

    def get_file_save_path(self):
        unique_path = secrets.token_hex(15) # see notes
        path = self.program.get_file_saving_path() + 'supplier_inquiries/emails/' + unique_path
        return path

    def get_program_version_ids(self):
        order_inquiries = self.program.orderinquiry_set.all()
        version_ids = []
        supplier_selection_complete = True
        for order_inquiry in order_inquiries:
            version = order_inquiry.get_current_costing_version()
            if version:
                version_ids.append(version.pk)

            # If a single version of the supplier inquiries are incomplete skip sending emails for program
            if not version or not version.supplier_inquiries_complete:
                supplier_selection_complete = False
                break

        # If supplier selection is not complete even for One, set version_ids to empty list.
        if not supplier_selection_complete:
            version_ids = []

        return version_ids

    def get_pending_suppliers(self):
        supplier_ids = SupplierInquiry.objects.filter(version_id__in=self.program_version_ids).values_list('supplier_id', flat=True).distinct('supplier')
        suppliers = Supplier.objects.filter(pk__in=supplier_ids)  # get unique suppliers of order_id
        return suppliers

    def get_pending_supplier_inquiries(self, supplier):
        email_pending_supplier_inquiries = SupplierInquiry.objects.filter(
            version_id__in=self.program_version_ids,
            supplier=supplier,
            email_status=SupplierInquiry.QUEUED_EMAIL,
            # version__supplier_inquiries_complete=True
        )
        return email_pending_supplier_inquiries

    def get_cc_emails(self):
        order_inquiries = self.program.orderinquiry_set.all()
        cc_emails = []

        for order_inquiry in order_inquiries:
            if order_inquiry.merchant:
                cc_emails.append(order_inquiry.merchant.email)

            for watcher in order_inquiry.watchlist.all():
                cc_emails.append(watcher.email)
        return cc_emails

    def get_customer_brand(self):
        data = {
            'customer': self.program.customer.name,
            'brand': self.program.brand.name,
        }
        return data

    def send_supplier_inquiry_email(self):
        suppliers = self.get_pending_suppliers()
        # print(suppliers)
        if not suppliers.exists():
            return {'status': 'No Iquiries to Send Email'}

        # Send only if it is not part of a program
        self.process_emails(suppliers)


class ConsolidateSupplierInquiryEmailSender(InquiryEmailSender):

    def __init__(self):
        self.version = None

    def get_file_saving_path(self):
        return 'consolodate/'

    def get_file_save_path(self):
        unique_path = secrets.token_hex(15) # see notes
        path = self.get_file_saving_path() + 'supplier_inquiries/emails/' + unique_path
        return path

    def get_pending_suppliers(self):
        supplier_ids = SupplierInquiry.objects.filter(version=None).values_list('supplier_id', flat=True).distinct('supplier')
        suppliers = Supplier.objects.filter(pk__in=supplier_ids)  # get unique suppliers of order_id
        return suppliers

    def get_pending_supplier_inquiries(self, supplier):
        email_pending_supplier_inquiries = SupplierInquiry.objects.filter(
            version=None,
            supplier=supplier, email_status=SupplierInquiry.QUEUED_EMAIL
        )
        return email_pending_supplier_inquiries

    def get_cc_emails(self):
        cc = None
        return cc

    def get_customer_brand(self):
        data = {}
        if self.version:
            data = {
                'customer': self.version.order.customer.name,
                'brand': self.version.order.brand.name,
            }
        return data

    def send_supplier_inquiry_email(self):

        suppliers = self.get_pending_suppliers()
        if not suppliers.exists():
            return {'status': 'No Iquiries to Send Email'}
        self.process_emails(suppliers)


class ProcessSupplierInquiryEmail:
    highlighted_columns = InquiryEmailSender.highlighted_columns

    def get_unit_mappings(self):

        mappings = {
            MaterialUnitHelper.INCHES_UNIT: [
                'in', 'inches', 'inch'
            ],
            MaterialUnitHelper.METERS_UNIT: [
                'm', 'mtrs', 'mtr', 'meters', 'meter', 'mtr'
            ],
            MaterialUnitHelper.CENTIMETERS_UNIT: [
                'cm', 'centimeters', 'centimetres', 'centimeter', 'centimetre',
            ],
            MaterialUnitHelper.MILLIMETERS_UNIT: [
                'mm', 'millimeters', 'millimetres', 'millimeter', 'millimetre',
            ],
            MaterialUnitHelper.YARDS_UNIT: [
                'yard', 'yards', 'yrd'
            ],
            MaterialUnitHelper.FEETS_UNIT: [
                'feet', 'foot', 'feets'
            ],
        }
        return mappings

    def get_mapped_width_unit(self, quote_units):
        # print("Aruna", [quote_units.lower()])
        mappings = self.get_unit_mappings()
        mapped_unit = ''
        
        for unit, unit_mappings in mappings.items():
            # print(unit_mappings)
            if quote_units.lower() in unit_mappings:
                mapped_unit = quote_units
                break
        return mapped_unit

    def process_supplier_quote_input(self, field, value):
        processed_value = None
        float_regex = '\d+\.?\d+'
        # print(field, re.search(float_regex, str(value),0))
        if value:
            if field in ['fob_price', 'cif_price', 'ex_work_price', 'lead_time', 'minimum_order_quantity']:
                try:
                    # print(field, re.search(float_regex, str(value)))
                    processed_value = value #re.search(float_regex, str(value)).group(0)
                    # print("test")
                except IndexError:
                    pass
            elif field == 'cutting_width':
                try:
                    processed_value = value #re.search(float_regex, str(value)).group(0)
                except IndexError:
                    pass
            elif field in ['ship_mode', 'pay_mode', 'costing_unit', 'cutting_width_unit', 'minimum_order_quantity_units']:
                processed_value = choice_to_reverse_dictionary(SupplierInquiry().get_lists_of_validation_for_excel()[1][field]).get(value, None)
        return processed_value

    def __init__(self, email):
        self.email = email

    def get_supplier_inquiries(self, row):
        td_values = []
        for data in row:
            if row[data]:
                # print(row[data])
                for splitted_data in str(row[data]).split(","):
                    td_values.append(splitted_data)
        # td_values = [splitted_data for data in row if not row[data] == None for splitted_data in str(row[data]).split(",") ]
        # print(td_values)
        supplier_inquiries = SupplierInquiry.objects.filter(hash_id__in=td_values)

        return supplier_inquiries

    def update_inquiries(self, inquiry, data):
        inquiry_details = get_object_or_none(SupplierInquiryDetail,{'supplier_inquiry':inquiry.id, 'completed': False})
        # print(inquiry_details)
        # print(data)
        if inquiry_details:
            ritz_code = ''
            supplier_reference_code = ''
            for field, field_value in data.items():
                if field == 'supplier_reference_code':
                    supplier_reference_code = field_value
                processed_value = self.process_supplier_quote_input(field, field_value)
                # print(processed_value,field, field_value)
                if field_value != None:
                    
                    setattr(inquiry_details, field,processed_value)
            if not supplier_reference_code == '':
                supplier_reference_code_object, created = SupplierInquiryMaterialCode.objects.get_or_create(customer_brand_material = inquiry_details.supplier_inquiry.customer_brand_material,
                                                                  supplier_material_reference_code = supplier_reference_code,
                                                                  supplier=inquiry.supplier)
                if supplier_reference_code_object:
                    inquiry_details.supplier_inquiry_material_code = supplier_reference_code_object
            inquiry_details.save()
            inquiry.email_status = inquiry.RECEIVED_AND_PROCESSED
            inquiry.save()
            # print("saved")


    def process_data(self, inquiries, row_data):
        inquiry_detail_data = {}
        first_inquiry = inquiries.first()

        highlighted_columns = OrderVersionSupplierInquiryEmailSender.highlighted_columns
        headers = []
        if first_inquiry.customer_brand_material:
            headers = OrderVersionSupplierInquiryEmailSender.get_material_headers(first_inquiry.customer_brand_material.material_type)
        elif first_inquiry.item_service:
            headers = OrderVersionSupplierInquiryEmailSender.get_service_headers(first_inquiry.item_service.service_type)
        # headers.append({'label': 'Cost Per Unit', 'name': 'cost_per_unit'})
        headers.append({'label': 'Costing Unit', 'name': 'costing_unit'})
        # print(row_data)
        # return
        if headers:
            index = 0
            for header in headers:
                # print(header)
                # if header['name'] in highlighted_columns:
                    # print(row_data[index])
                if header['label'] in row_data:
                    if 'name' in header:
                        inquiry_detail_data[header['name']] = row_data[header['label']]
            # print(inquiry_detail_data)

        for supplier_inquiry in inquiries:
            self.update_inquiries(supplier_inquiry, inquiry_detail_data)

    def get_last_header_column(self,row, current_column):
        last_header_column = current_column
        for cell in row[current_column:]:
            if cell.value == None:
                break
            last_header_column+=1
        if not last_header_column == current_column:
            last_header_column-=1
        return last_header_column
    
    def get_next_last_none_empty_row(self,sheet,current_row):
        next_last_none_empty_row = current_row
        for row in range(current_row,sheet.max_row+1):
            # if all([cell.value == None for cell in sheet[row]]):
            if sheet.cell(row = row, column = 2).value == None:
                break
            next_last_none_empty_row+=1
        if not next_last_none_empty_row == current_row:
            next_last_none_empty_row-=1
        return next_last_none_empty_row

    def get_tables_from_excel(self,excel_file):
        
        tables = []
        sheet = excel_file.active
        for r in range(1,sheet.max_row+1):
            material_name = sheet.cell(row=r,column=1).value
            if not material_name == None:
                # print(material_name)
                r+=1
                if not all([cell.value == None for cell in sheet[r]]):
                    table_header_row = r
                    last_table_row = self.get_next_last_none_empty_row(sheet,r+1)
                    last_header_column = self.get_last_header_column(sheet[table_header_row],2)
                    # print("table", last_table_row)
                    for data_row in range(r+1,last_table_row+1):
                        # print(data_row)
                        table_row_data = {}
                        for col in range(2,last_header_column+2):
                            table_header = sheet.cell(row=table_header_row, column=col).value
                            # print(table_header)
                            table_row_data[table_header]=sheet.cell(row=data_row, column=col).value
                        tables.append(table_row_data)
        return tables
    def process_email(self):
        # email_body = self.email.body#.replace("=\r\n", '').replace('\r', '').replace('\n', '')
        # print(email_body)
        # return
        # soup = BeautifulSoup(email_body, "html.parser")
        attachments = self.email.attachments.all()
        for attachment in attachments:
            if attachment.type == '.xlsx':
                # excel_file = openpyxl.open('./attachments/Price Inquiry For RITZ Clothing.xlsx')
                try:
                    excel_file = openpyxl.load_workbook(handle_file_read(attachment.file_path))
                except:
                    continue
                tables = self.get_tables_from_excel(excel_file)
                # print(tables)
                # print(tables)
                for row_data in tables:
                    # print(row_data)
                    supplier_inquiries = self.get_supplier_inquiries(row_data)
                    # print(supplier_inquiries)
                    if supplier_inquiries.exists():
                        current_inquiries = supplier_inquiries
                        inquiries_updated = 0
                        # print(row_data)
                        first_inquiry = current_inquiries.first() if current_inquiries else None
                        if inquiries_updated < 3 and first_inquiry and first_inquiry.customer_brand_material and first_inquiry.customer_brand_material.material_type == Material.FABRIC_MATERIAL:
                            self.process_data(current_inquiries, row_data)

                        elif first_inquiry and first_inquiry.customer_brand_material and first_inquiry.customer_brand_material.material_type != Material.FABRIC_MATERIAL:
                            self.process_data(current_inquiries, row_data)
                            current_inquiries = None

                        else:
                            if first_inquiry:
                                self.process_data(current_inquiries, row_data)
                            current_inquiries = None
                    # print(supplier_inquiries)
        return
        # print(tables)
        # for table in tables:
        #     table_id = table.get('id')
        #     # print(table)
        #     if table_id and 'quote_data_table' in table_id: #or True:
        #         print("test2")
        #         current_inquiries = None
        #         inquiries_updated = 0
        #         # print(table)
        #         table_rows = table.find_all('tr')
        #         # print(table_rows)
        #         for row in table_rows:
        #             # print(row)
        #             tds = row.find_all('td')
        #             # print(tds)
        #             td_index = 0
        #             row_data = []
        #             for td in tds:
        #                 td_value = ' '.join(td.renderContents().decode("utf-8").split()).strip()
        #                 # print(td_value)
        #                 # break
        #                 row_data.append(td_value)
        #                 if td_index == 0:
        #                     supplier_inquiries = self.get_supplier_inquiries(td_value)
        #                     # print(supplier_inquiries)
        #                     if supplier_inquiries.exists():
        #                         current_inquiries = supplier_inquiries
        #                         inquiries_updated = 0

        #                 td_index += 1
        #             # print(row_data)

        #             first_inquiry = current_inquiries.first() if current_inquiries else None
        #             if inquiries_updated < 3 and first_inquiry and first_inquiry.customer_brand_material and first_inquiry.customer_brand_material.material_type == Material.FABRIC_MATERIAL:
        #                 self.process_data(current_inquiries, row_data)

        #             elif first_inquiry and first_inquiry.customer_brand_material and first_inquiry.customer_brand_material.material_type != Material.FABRIC_MATERIAL:
        #                 self.process_data(current_inquiries, row_data) # issue detected of row_data
        #                 current_inquiries = None

        #             else:
        #                 if first_inquiry:
        #                     self.process_data(current_inquiries, row_data)
        #                 current_inquiries = None


