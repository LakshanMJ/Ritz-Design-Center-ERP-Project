import openpyxl
from django.conf import settings
from django.db import transaction
import os
from shared.models import Supplier, Address, SupplierLocation, SupplierContactPerson, SupplierBrand, CustomerBrand
import re


class SuppplierListProcessor:

    _brand_text = ['Brand Name', 'Brand', 'brand']
    _supplier_text = ['Supplier Name', 'Supplier', 'supplier']
    _location_text = ['Local/Foreign', 'Local / Foreign', 'local/foreign']
    _type_text = ['Type', 'type']
    
    def extract_int_from_string(self, value):
        try:
            integers = [int(match) for match in re.findall(r'\d+', value)]
            return integers[0] if integers else 0
        except ValueError:
            return 0 
        
    field_mapping = {
        'Brand Name': 'brand_name',
        'Supplier Name': 'name',
        'Local/Foreign': 'location',
        'Raw Meterial': 'raw_material',
        'Service': 'service',
        'Supplier Address': 'supplier_location',
        'Item Description': 'item_description',
        'Contact person 1': 'contact_person_1',
        'E-mail address 1': 'email_address_1',
        'Contact person 2': 'contact_person_2',
        'E-mail address 2': 'email_address_2',
        'Payment Term': 'payment_term',
        'Shipping mode': 'shipping_mode',
        'Ex.fty to Inhouse': 'ex_fty_to_inhouse',
        'FOB to inhouse': 'fob_to_inhouse',
    }

    def process_columns(self, sheet):
        errors = []
        header_row_index = 0
        for row_index in range(1, sheet.max_row + 1):
            
            for column_index in range(1, sheet.max_column + 1):
                brand_cell_value = sheet.cell(row=row_index, column=column_index).value
                supplier_cell_value = sheet.cell(row=row_index, column=column_index+1).value
                location_cell_value = sheet.cell(row=row_index, column=column_index+2).value

                if brand_cell_value in self._brand_text and supplier_cell_value in self._supplier_text and location_cell_value in self._location_text:
                    header_row_index = row_index
                    first_headers = [cell.value for cell in sheet[row_index]]
                    first_header_indices = {header: idx for idx, header in enumerate(first_headers)}
                    second_headers = [cell.value for cell in sheet[row_index+1]]
                    second_header_indices = {header: idx for idx, header in enumerate(second_headers)}
                else:
                    errors.append("Supplier list format is not valid. Unable to locate headers. Please make sure the row contains the headers")
        return errors, header_row_index, first_header_indices, second_header_indices

    def read_supplier_data(self):
        file_path = "shared/dataimports/data_files/supplier_list.xlsx"

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        errors, header_row_index, first_header_indices, second_header_indices = self.process_columns(sheet)
        
        row_data = []
        for row_index in range(header_row_index+2, sheet.max_row + 1):
            row_dict = {}
            for header, col_index in first_header_indices.items():
                if header is not None and header not in self._type_text:
                    cell_value = sheet.cell(row=row_index, column=col_index+1).value
                    row_dict[header] = cell_value
            for header, col_index in second_header_indices.items():
                if header is not None:
                    cell_value = sheet.cell(row=row_index, column=col_index+1).value
                    row_dict[header] = cell_value
            row_data.append(row_dict)
        return row_data

    def create_suppliers(self):
        errors = []
        data = self.read_supplier_data()
        suppliers = []
        for entry in data:
            supplier = Supplier()
            address = Address()
            location = SupplierLocation()
            contact_person_1 = SupplierContactPerson()
            contact_person_2 = SupplierContactPerson()
            for key, value in entry.items():
                if key in self.field_mapping:
                    field_value = self.field_mapping[key]
                    if field_value == 'brand_name':
                        customer_brands = CustomerBrand.objects.filter(brand__name=value)
                    if field_value == 'name':
                        supplier.name = value
                    if field_value == 'location':
                        supplier.location = value
                    supplier.raw_material = True if field_value == 'raw_material' and value == 'Yes' else False
                    supplier.service = True if field_value == 'service' and value == 'Yes' else False
                    
                    if field_value == 'ex_fty_to_inhouse' and value is not None:
                        supplier.ex_fty_to_inhouse =  self.extract_int_from_string(value)   
                    else:
                        supplier.ex_fty_to_inhouse = 0
                    if field_value == 'fob_to_inhouse' and value is not None:
                        supplier.fob_to_inhouse =  self.extract_int_from_string(value)   
                    else:
                        supplier.fob_to_inhouse = 0
                    if field_value == 'supplier_location':
                            if value:
                                address_lines = value.split('\n')
                                if len(address_lines) >= 3:
                                    address.address_line_1 = address_lines[0]
                                    address.address_line_2 = address_lines[1] + ', ' + address_lines[2]
                                else:
                                    if address_lines:
                                        full_address = ''
                                        for address_line in address_lines:
                                            full_address += address_line
                                            address.address_line_1 = full_address
                    if field_value == 'payment_term':
                        if value == Supplier.TT_IN_ADVANCE_100_PRECENT_PAYMENT_METHOD_TYPE:
                            supplier.payment_term = Supplier.TT_IN_ADVANCE_100_PRECENT_PAYMENT_METHOD_TYPE
                        elif value == Supplier.TT_ADVANCE_PAYMENT_METHOD_TYPE:
                            supplier.payment_term = Supplier.TT_ADVANCE_PAYMENT_METHOD_TYPE
                        elif value == Supplier.CREDIT_DAYS_30_PAYMENT_METHOD_TYPE:
                            supplier.payment_term = Supplier.CREDIT_DAYS_30_PAYMENT_METHOD_TYPE
                        elif value == Supplier.CREDIT_DAYS_140_PAYMENT_METHOD_TYPE:
                            supplier.payment_term = Supplier.CREDIT_DAYS_140_PAYMENT_METHOD_TYPE
                        elif value == Supplier.LC_PAYMENT_METHOD_TYPE:
                            supplier.payment_term = Supplier.LC_PAYMENT_METHOD_TYPE
                        else:
                            supplier.payment_term = None
                    if field_value == 'contact_person_1':
                        contact_person_1.name = value
                    if field_value == 'email_address_1':
                        contact_person_1.email = value
                    if field_value == 'contact_person_2':
                        contact_person_2.name = value
                    if field_value == 'email_address_2':
                        contact_person_2.email = value
                     #setattr(supplier, field_value, value)
            suppliers.append(supplier)
            
            if supplier.name:
                supplier.save()
                contact_person_1.supplier = supplier
                contact_person_2.supplier = supplier
                contact_person_1.save()
                contact_person_2.save()

                address.save()
                location.supplier = supplier
                location.address = address
                supplier.email = 'tharindu.s@ritzclothing.lk'
                location.save()
                supplier.supplier_location = address
                for customer_brand in customer_brands:
                    supplier.customer_brand.add(customer_brand)
                supplier.save()
        return errors
