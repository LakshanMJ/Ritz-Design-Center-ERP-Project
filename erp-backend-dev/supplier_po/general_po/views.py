from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.views import APIView
from django.core.exceptions import ObjectDoesNotExist

from marketing.mixins.order_mixins import OrderMixin
from marketing.serializers import PurchaseOrderBomBasicSerializer
from materials.fieldmetadata.measuring_unit_helpers import MaterialUnitHelper
from materials.scripts.supplier_po_bom_generator import NewGenerateSupplierPOs
from materials.serializers.material_serializers import CustomerBrandMaterialBasicSerializer
from shared.helpers.general_results_set_pagination import GeneralLargeResultsSetPagination
from shared.permissions.view_permissions import HasPermission
from shared.permissions.roles import MERCHANT_ROLE, ADMIN_ROLE
from rest_framework.generics import get_object_or_404
from marketing.models import OrderCostingVersion, OrderPack, GeneralPO, GeneralPOQuantity, GeneralPOMaterialQuantity, \
    SupplierDeliveryDate, SupplierDeliveryDateQuantity, \
    POPackItemPlacement, POPackPlacement, SupplierInquiry, SupplierInquiryDetail, PurchaseOrderBom, ActualPOClub, FabricGRNDetail
from supplier_po.general_po.serializers import GeneralPOSerializer, GeneralPOMaterialQuantitySerializer, \
    GeneralPOSupplierMaterialPriceSerializer, GeneralPOMaterialQuantityDetailSerializer, GeneralPOShadeSummarySerializer
from supplier_po.helpers.supplier_po_bom_generator import SupplierPOClubBOMGenerator
from supplier_po.models import SupplierDeliveryDateQuantityPOAllocation, GeneralPOSupplierMaterialPrice, SupplierPO, GeneralPOSupplier, \
    SupplierPOGRNMaterial
from supplier_po.supplier_po.serializers import SupplierDeliveryDateQuantitySerializer, SupplierPOSerializer, SupplierDeliveryDateQuantityListSerializer
from supplier_po.supplier_po_grn.serializers import GRNShadeSummarySerializer
from supplier_po.general_po.serializers import GeneralPOFabricChartSummarySerializer, SupplierDeliveryDateQuantityPOAllocationSerializer, GeneralPOMaterialQuantityListSerializer
from shared.serializers import SupplierSerializer
from shared.models import Supplier, Plant, Port, AIR_TRANSPORT_METHOD, ROAD_TRANSPORT_METHOD, SEA_TRANSPORT_METHOD
from materials.models import CustomerBrandMaterial, SupplierInquiryMaterialCode
from shared.utils import clean_search_dictionary, get_object_or_none, clean_search_dictionary, \
    search_qs_from_global_filter, get_object_or_none_dict
from materials.serializers.material_serializers import SupplierInquiryMaterialDetailSerializer
from marketing.serializers import PurchaseOrderBasicSerializer
from materials.models import Material, SEWING_TRIM_TYPES, PACKAGING_TYPES
from marketing.permissions.costing_permissions import ObjectStatePermissionMixin
from shared.utils import round_if_number, ceil_number, convert_per_unit_cost


class GeneralPOMetaDataDetailView(APIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]

    def get_states(self):
        genera_po_states = []
        for choice in GeneralPO.STATE_CHOICES:
            genera_po_states.append({"id": choice[0], "name": choice[1]})
        return genera_po_states
    
    def get(self, request, *args, **kwargs):
        data = {
            'states': self.get_states()
        }
        return Response(data, status=status.HTTP_200_OK)
    

class GeneralPOOrderPackQuantityDetailView(APIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]
    
    def get(self, request, *args, **kwargs):
        data = []
        id = kwargs.get('general_po_id', None)
        general_po = get_object_or_404(GeneralPO, id=id)
        quantities = general_po.get_quantities()
        for quantity in quantities:
            data.append({
                'id':quantity.id,
                'country': quantity.pack.country.id,
                'colorway': quantity.pack.colorway.id,
                'size': quantity.pack.size.id,
                'pack_id': quantity.pack.id,
                'quantity': quantity.quantity
            })

        return Response(data, status=status.HTTP_200_OK)
    

class GeneralPOOrderPackQuantitySaveView(APIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]
    
    def validate_data(self, instance):
        pass

    def post(self, request, *args, **kwargs):
        id = kwargs.get('costing_version_id', None)
        costing = get_object_or_404(OrderCostingVersion, id=id)
        quantities = request.data.get('quantities', [])

        general_po = GeneralPO.objects.create(
            costing=costing,
            state=GeneralPO.DRAFT
        )

        for row in quantities:
            country_id = row['id']
            colorways = row['colorways']
            for colorway in colorways:
                colorway_id = colorway['id']
                for size_group in colorway['size_groups']:
                    for size in size_group['sizes']:
                        size_id = size['id']
                        quantity = size['quantity']
                        pack = OrderPack.objects.get(size=size_id, country=country_id, colorway=colorway_id, version=costing)
                        general_po_quantity, created = GeneralPOQuantity.objects.get_or_create(
                            general_po=general_po,
                            pack=pack
                        )
                        general_po_quantity.quantity = quantity
                        general_po_quantity.save()
        return Response({'general_po_id': general_po.id, 'status': True}, status=status.HTTP_200_OK)
    

class GeneralPOOrderPackQuantityUpdateView(ObjectStatePermissionMixin, APIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]
    editable_states = [GeneralPO.DRAFT, GeneralPO.QUANTITY_VERIFICATION, ]

    def get_object_current_state(self):
        id = self.kwargs.get('general_po_id', None)
        general_po = get_object_or_404(GeneralPO, id=id)
        return general_po.state

    def handle_quantity_change(self, general_po, quantities_changed):
        from supplier_po.helpers.general_po_processor import GeneralPOBOM
        if quantities_changed:
            general_po_quantities = GeneralPOMaterialQuantity.objects.filter(
                general_po=general_po
            )
            general_po_bom = GeneralPOBOM(general_po)
            general_po_bom.create_general_po_bom()
            for general_po_quantity in general_po_quantities:
                general_po_quantity.order_quantity_discrepancy_reason = None
                general_po_quantity.order_quantity_discrepancy_other = None
                general_po_quantity.completed = False
                general_po_quantity.save()      
    
    def post(self, request, *args, **kwargs):
        id = kwargs.get('general_po_id', None)
        general_po = get_object_or_404(GeneralPO, id=id)
        quantities = request.data.get('quantities', [])
        quantities_changed = request.data.get('quantities_changed')

        self.handle_quantity_change(general_po, quantities_changed)

        for row in quantities:
            country_id = row['id']
            colorways = row['colorways']
            for colorway in colorways:
                colorway_id = colorway['id']
                for size_group in colorway['size_groups']:
                    for size in size_group['sizes']:
                        size_id = size['id']
                        quantity = size['quantity']
                        pack = OrderPack.objects.get(size=size_id, country=country_id, colorway=colorway_id, version=general_po.costing)
                        general_po_quantity, created = GeneralPOQuantity.objects.get_or_create(
                            general_po=general_po,
                            pack=pack
                        )
                        general_po_quantity.quantity = quantity
                        general_po_quantity.save()
        return Response({'general_po_id': general_po.id, 'status': True}, status=status.HTTP_200_OK)
    

class GeneralPOListView(generics.ListAPIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]
    serializer_class = GeneralPOSerializer
    queryset = GeneralPO.objects.all().order_by('-id')
    pagination_class = GeneralLargeResultsSetPagination

    sort_keys = {
        'id': 'id',
        'po_club': 'po_club_id',
        'costing': 'costing_id',
        'state': 'state',
    }

    def clean_dictionary(self, dic):
        replace_keys = {
            'id': 'id',
            'po_club': 'po_club_id',
            'costing': 'costing_id',
            'state': 'state',
        }
        dictionary = clean_search_dictionary(dic, replace_keys)
        return dictionary

    def get_queryset(self):
        qs = super().get_queryset()
        search_data = self.request.GET.dict()
        search_dictionary = self.clean_dictionary(search_data)
        global_filter = self.request.query_params.get('global_filter', None)
        sort_col = self.request.query_params.get('sort_col', None)
        sort_col = self.sort_keys.get(sort_col, 'id')
        sort_dir = self.request.query_params.get('sort_dir', default='desc')
        search_fields = ['po_club_id', 'costing_id', 'state', 'id']
        qs = search_qs_from_global_filter(qs, global_filter, search_dictionary, search_fields, sort_col, sort_dir, GeneralPO)
        return qs

    
class GeneralPODetailView(APIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]

    def get_general_po_quantity(self, general_po, country, colorway, sizes):
        from django.db.models import Sum
        size_ids = sizes.filter().values_list('id', flat=True)
        quantity = 0
        total_ratio = OrderPack.objects.filter(country=country, colorway=colorway, size__in=size_ids, version=general_po.costing).aggregate(Sum('cad_quantity'))
        total_quantity = general_po.generalpoquantity_set.filter(
            pack__country=country,
            pack__colorway=colorway,
            pack__size__in=size_ids
        ).aggregate(Sum('quantity'))
        if total_quantity['quantity__sum']:
            quantity = round(total_quantity['quantity__sum'])
        return quantity
    
    def get_quantity(self, general_po, country, colorway, size):
        from marketing.models import GeneralPOQuantity
        quantity = 0
        pack = OrderPack.objects.get(country=country, colorway=colorway, size=size, version=general_po.costing)
        general_po_quantity = get_object_or_none(GeneralPOQuantity, 
            {
                'general_po': general_po,
                'pack': pack
            }
        )
        if general_po_quantity:
            quantity = general_po_quantity.quantity
        return quantity
    
    def get(self, request, *args, **kwargs):
        id = kwargs.get('general_po_id', None)
        general_po = get_object_or_404(GeneralPO, id=id)
        costing = general_po.costing
        is_club_po = general_po.is_po_club_general_po()
        po_club_id = None
        po_club_display_number = None
        purchase_orders = None
        if is_club_po:
            po_club_id = general_po.po_club.id
            po_club_display_number = general_po.po_club.display_number
            purchase_orders = general_po.po_club.get_purchase_orders()

        data = {
            'id': general_po.id,
            'display_number': general_po.display_number,
            'state': general_po.state,
            'state_display': general_po.get_state_display(),
            'costing_display_number': general_po.costing.display_number,
            'costing_id': general_po.costing.id,
            'order_display_number': general_po.costing.order.display_number,
            'order_id': general_po.costing.order.id,
            'plant_id': general_po.plant.id if general_po.plant else None,
            'palnt_name': general_po.plant.name if general_po.plant else None,
            'quantities': [],
            'size_groups': [],
            'is_club_po': is_club_po,
            'po_club_id' : po_club_id,
            'po_club_display_number': po_club_display_number,
            'purchaseorder_set': PurchaseOrderBasicSerializer(purchase_orders, many=True).data
        }
        
        order = general_po.costing.order
        countries = order.get_order_countries()
        colorways = order.get_order_colorways()
        for country in countries:
            country_data = {
                'id': country.id,
                'country_name': country.country.name,
                'colorways': []
            }
            for colorway in colorways:
                colorway_data = {
                    'id': colorway.id,
                    'colorway': colorway.colorway,
                    'size_groups': []
                }
                size_groups = order.get_order_size_groups()
                for group in size_groups:
                    sizes = group.sizes.all()
                    group_data = {
                        'id': group.id,
                        'sizes': [],
                        'total_quantity': self.get_general_po_quantity(general_po, country, colorway, sizes)
                    }
                
                    for size in sizes:
                        size_data = {
                            'id': size.id,
                            'name': size.size.name,
                            'quantity': self.get_quantity(general_po, country, colorway, size)
                        }
                        group_data['sizes'].append(size_data)
                    colorway_data['size_groups'].append(group_data)
                country_data['colorways'].append(colorway_data)
            data['quantities'].append(country_data)

        size_groups = order.get_order_size_groups()
        for group in size_groups:
            group_data = {
                'id': group.id,
                'sizes': []
            }
            sizes = group.sizes.all()
            for size in sizes:
                size_data = {
                    'id': size.id,
                    'name': size.size.name
                }
                group_data['sizes'].append(size_data)
            data['size_groups'].append(group_data)
        
        return Response(data, status=status.HTTP_200_OK)
        

class GeneralPOStateUpdateView(APIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]
    
    def post(self, request, *args, **kwargs):
        id = kwargs.get('general_po_id', None)
        general_po = get_object_or_404(GeneralPO, id=id)
        new_state = request.data.get('new_state', None)
        plant_id = request.data.get('plant_id', None)
        modal = request.data.get('modal', False)
        user = request.user

        if plant_id:
            plant = get_object_or_404(Plant, pk=plant_id)
            general_po.plant=plant
            general_po.save()

        if modal:
            user = self.request.user
            has_admin_role = user.has_role(ADMIN_ROLE)
            if user.is_authenticated and has_admin_role:
                errors = general_po.move_to_next_state(new_state, user)
                if not errors:
                    http_response = Response({'status': True})
                else:
                    http_response = Response({'error': errors}, status=status.HTTP_400_BAD_REQUEST)
            else:
                http_response = Response({'error': 'User not authenticated'}, status=status.HTTP_403_FORBIDDEN)
        else:
            errors = general_po.move_to_next_state(new_state, user)
            if not errors:
                http_response = Response({'status': True})
            else:
                http_response = Response({'error': errors}, status=status.HTTP_400_BAD_REQUEST)
        return http_response
    

class GeneralPOBOMListViewView(generics.ListAPIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]
    serializer_class = GeneralPOMaterialQuantitySerializer
    queryset = GeneralPOMaterialQuantity.objects.all()

    def get_queryset(self):
        qs = super().get_queryset().filter(
            general_po=self.kwargs['general_po_id']
        ).order_by('material__material_detail__generic_material__user_material__display_order')
        return qs
    

class GeneralPOMaterialQuantitySendPOForMaterialStateUpdateView(generics.ListAPIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]

    def post(self, request):
        data = request.data.get('material_quantity_ids', [])
        for row in data:
            send_po_for_material = row['send_po_for_material'] 
            material_quantity = get_object_or_404(GeneralPOMaterialQuantity, pk=row['id'])
            material_quantity.send_po_for_material = send_po_for_material
            material_quantity.save()
        return Response({'success': True})


class OrderInquirySupplierInquiryList(APIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]

    def get_costing_materials(self, order, material):
        item_material_ids = POPackItemPlacement.objects.filter(
            po_material=material, po_pack_item__order_pack_item__item__order=order
        ).values_list('costing_pack_item_placement__orderpackitemplacementmaterial__material_id', flat=True).distinct()
        pack_material_ids = POPackPlacement.objects.filter(
            po_material=material, po_pack__order_pack__size__order=order
        ).values_list('costing_pack_placement__orderpackplacementmaterial__material_id', flat=True).distinct()
        material_ids = [*item_material_ids, *pack_material_ids]
        return material_ids

    def get(self, request, *args, **kwargs):
        data = []
        general_po_material_quantity = get_object_or_404(GeneralPOMaterialQuantity, pk=self.kwargs.get('general_po_material_quantity_id'))
        material = get_object_or_404(CustomerBrandMaterial, pk=self.kwargs.get('material_id'))
        material_ids = self.get_costing_materials(general_po_material_quantity.general_po.costing.order, material)
        versions = general_po_material_quantity.general_po.costing.order.get_order_versions()
        exclude_ids = GeneralPOSupplierMaterialPrice.objects.filter(
            general_po_supplier__general_po=general_po_material_quantity.general_po
        ).values_list('supplier_inquiry_detail', flat=True)
        supplier_inquiry_details = SupplierInquiryDetail.objects.filter(
            supplier_inquiry__customer_brand_material=material, 
            supplier_inquiry__version__in=versions
        ).exclude(
            id__in=exclude_ids
        )
        print(supplier_inquiry_details, exclude_ids)
        suppliers = Supplier.objects.filter(id__in=supplier_inquiry_details.values_list('supplier_inquiry__supplier', flat=True))
        for supplier in suppliers:
            inquiries = supplier_inquiry_details.filter(supplier_inquiry__supplier=supplier)
            data.append({
                'supplier_id': supplier.id,
                'supplier_name': supplier.name,
                'inquries': SupplierInquiryMaterialDetailSerializer(inquiries, many=True).data
            })

        return Response(data)
    

class GeneralPOBOMMaterialQuantityDetailView(generics.RetrieveAPIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]
    serializer_class = GeneralPOMaterialQuantitySerializer
    queryset = GeneralPOMaterialQuantity.objects.all()


class GeneralPOBOMSupplierDeliveryDateCreateUpdateView(APIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]

    def get_valid_order_inquiry_detail(self, costing, supplier, material):
        supplier_inquiry_detail = get_object_or_none(
            SupplierInquiryDetail, 
            {
                'supplier_inquiry__customer_brand_material':material,
                'supplier_inquiry__version': costing,
                'supplier_inquiry__supplier': supplier
            }
        )
        return supplier_inquiry_detail
    
    def post(self, request, *args, **kwargs):
        id = kwargs.get('general_po_material_quantity_id', None)
        general_po_material_quantity = get_object_or_404(GeneralPOMaterialQuantity, id=id)
        costing = general_po_material_quantity.general_po.costing
        material = general_po_material_quantity.material
        quantities = request.data.get('purchase_order_club_bom_suppliers', [])
        order_quantity = request.data.get('order_quantity', None)
        order_quantity_units = request.data.get('order_quantity_units', None)
        general_po_material_quantity.order_quantity = order_quantity
        general_po_material_quantity.order_quantity_units = order_quantity_units
        general_po_material_quantity.save()

        for row in quantities:
            purchase_order_bom_supplier_id = row['id']
            quantity = row['quantity']
            supplier_id = row['supplier']
            delivery_date = row['confirmed_delivery_date']
            supplier = Supplier.objects.get(id=supplier_id)

            supplier_inquiry_detail = self.get_valid_order_inquiry_detail(costing, supplier, material)
            supplier_delivery_date, created = SupplierDeliveryDate.objects.get_or_create(
                general_po=general_po_material_quantity.general_po,
                supplier=supplier,
                confirmed_delivery_date=delivery_date
            )
            if purchase_order_bom_supplier_id:
                purchase_order_bom_supplier = SupplierDeliveryDateQuantity.objects.get(id=purchase_order_bom_supplier_id)
                purchase_order_bom_supplier.supplier=supplier
                purchase_order_bom_supplier.supplier_inquiry_detail=supplier_inquiry_detail
                purchase_order_bom_supplier.quantity=quantity['quantity']
                purchase_order_bom_supplier.supplier_delivery_date=supplier_delivery_date
                purchase_order_bom_supplier.save()
            else:
                purchase_order_bom_supplier = SupplierDeliveryDateQuantity.objects.create(
                    supplier=supplier,
                    supplier_inquiry_detail=supplier_inquiry_detail,
                    quantity=quantity['quantity'],
                    quantity_units=general_po_material_quantity.measuring_unit,
                    supplier_delivery_date=supplier_delivery_date,
                    general_po_material_quantity=general_po_material_quantity
                )
        return Response({'status': True}, status=status.HTTP_200_OK)
    

class GeneralPOBOMSupplierDeleteView(generics.RetrieveDestroyAPIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]
    serializer_class = SupplierDeliveryDateQuantitySerializer
    queryset = SupplierDeliveryDateQuantity.objects.all()


class SaveSupplierQuantityDelivery(ObjectStatePermissionMixin, APIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]
    editable_states = [GeneralPO.DRAFT, GeneralPO.QUANTITY_VERIFICATION, ]

    ORDER_PRICE_ERROR_KEY = 'order_price'
    EXCESS_THRESHOLD_ERROR_KEY = 'excess_threshold'
    DELIVERY_QUANTITY_ERROR_KEY = 'quantity'
    DELIVERY_DATE_ERROR_KEY = 'confirmed_delivery_date'
    SHIP_MODE_KEY = 'transport_method'

    errors = {
    }

    def __init__(self):
        super().__init__()
        self.errors = {}

    def get_object_current_state(self):
        general_po_material_quntity = get_object_or_404(GeneralPOMaterialQuantity, pk=self.kwargs.get('general_po_material_quantity_id'))
        general_po = general_po_material_quntity.default_material_supplier.general_po_supplier.general_po
        return general_po.state

    def validate_po_bom_order_quantity(self, po_breakdown_data, suppliers):
        from decimal import Decimal
        errors = {}
        total_ordered_quantity = 0
        
        for row in po_breakdown_data['purchase_order_boms']:
            po_bom_id = row['id']
            order_quantity = row.get('order_quantity')
            quantity = row.get('quantity')

            if not order_quantity:
                errors[po_bom_id] = "Please enter order quantity."
            else:
                total_ordered_quantity += order_quantity

            # if quantity and order_quantity:
            #     if round(Decimal(quantity),2) != round(Decimal(order_quantity),2):
            #         errors[po_bom_id] = "Cannot mismatch allocated quantity and order quantity."

        
        total_delivery_quantity_error = self.validate_po_bom_total_delivery_quantity(total_ordered_quantity, suppliers)
        
        if total_delivery_quantity_error:
            errors['po_total_quantity_error'] = "Total required quantity and order quantity mismatch."

        if errors:
            self.errors['purchase_order_errors'] = errors

        return errors
    
    def validate_total_order_quantity(self, suppliers, material_total_order_quantity):
        total_delivery_quantity_error = self.validate_po_bom_total_delivery_quantity(material_total_order_quantity, suppliers)
        if total_delivery_quantity_error:
            self.errors['material_total_order_quantity_error'] = "Total required quantity and order quantity mismatch."

    def validate_po_bom_total_delivery_quantity(self, total_ordered_quantity, supplier_data):
        error = None
        total_deliveries_quantity = 0
        # for row in supplier_data:
        #     for delivery in row.get('deliveries', []):
        #         quantity = delivery.get('quantity', None)
        #         if quantity['quantity']:
        #             total_deliveries_quantity += quantity['quantity']
        # if total_deliveries_quantity and total_ordered_quantity:
        #     if total_deliveries_quantity != total_ordered_quantity:
        #         error = "Total delivery quantity and Order quantity mismatch."

        return error

    def validate_supplier_price(self, supplier_data):
        price_errors = {}
        main_index = 0
        for row in supplier_data:
            supplier_price_data = row.get('supplier_prices', {})
            ship_mode = supplier_price_data.get('transport_method', None)
            order_price = supplier_price_data['order_price']
            if not order_price or not ship_mode:
                if main_index not in price_errors:
                    price_errors[main_index] = {}
                if not order_price:
                    price_errors[main_index][self.ORDER_PRICE_ERROR_KEY] = 'Order price cannot be empty.'
                elif not ship_mode:
                    price_errors[main_index][self.SHIP_MODE_KEY] = 'Transport Method cannot be empty.'
            main_index += 1
        if price_errors:
            self.errors['supplier_price_errors'] = price_errors

    def validate_excess_threshold(self, supplier_data):
        excess_threshold_errors = {}
        main_index = 0
        for row in supplier_data:
            supplier_price_data = row.get('supplier_prices', {})
            excess_threshold = supplier_price_data['excess_threshold']
            if not excess_threshold:
                if main_index not in excess_threshold_errors:
                    excess_threshold_errors[main_index] = {
                        self.EXCESS_THRESHOLD_ERROR_KEY: 'Excess thershold cannot be empty.'
                    }
            main_index += 1
        if excess_threshold_errors:
            self.errors['supplier_price_errors'] = excess_threshold_errors

    def validate_delivery_errors(self, supplier_data, is_po_club_general_po):
        supplier_errors = {}
        main_index = 0
       
        for row in supplier_data:
            delivery_errors = {}
            delivery_index = 0
    
            for delivery in row.get('deliveries', []):
                errors = {}
                confirmed_delivery_date = delivery.get('confirmed_delivery_date')
                if not confirmed_delivery_date:
                    errors['confirmed_delivery_date'] = 'Select delivery date.'

                quantity = delivery.get('quantity', None)
                if not quantity['quantity']:
                    errors['quantity'] = 'Enter quantity.'

                port = delivery.get('port', None)
                transport_method = delivery.get('transport_method', None)
                port_object = get_object_or_none_dict(Port, pk=port)
                
                if transport_method in [AIR_TRANSPORT_METHOD, SEA_TRANSPORT_METHOD]:
                    if port and port_object.port_type != transport_method:
                        errors['port'] = f"Please select a valid port. Selected port is a {port_object.port_type}. Selected 'Transport Type' is '{transport_method}'"
                else:
                    if not transport_method:
                        errors['port'] = "Please select transport type."


                allocation_errors = {}
                allocation_index = 0
                total_allocation_quantity = 0
                allocations = delivery.get('supplierdeliverydatequantitypoallocation_set', [])

                if is_po_club_general_po:
                    if not allocations:
                        errors['purchase_order_set'] = 'Select at least single purchase order.'

                for allocation in allocations:
                    allocation_error = {}
                    purchase_order = allocation.get('purchase_order')
                    allocation_quantity = allocation.get('quantity')
                    
                    if not purchase_order:
                        allocation_error['purchase_order'] = 'Select purchase order.'
                    
                    if not allocation_quantity:
                        allocation_error['quantity'] = 'Enter quantity.'
                    else:
                        total_allocation_quantity += allocation_quantity

                    if allocation_error:
                        allocation_errors[allocation_index] = allocation_error
                    
                    allocation_index += 1 
                if allocation_errors:
                    errors['allocation_error'] = allocation_errors

                if quantity and total_allocation_quantity:
                    total_quantity_error = self.validate_delivery_total_quantity(quantity['quantity'], total_allocation_quantity)
                    if total_quantity_error:
                        errors['total_allocation_quantity_error'] = total_quantity_error

                if errors:
                    delivery_errors[delivery_index] = errors
                delivery_index += 1

            if delivery_errors:
                supplier_errors[main_index] = delivery_errors
                
            main_index += 1
            if supplier_errors:
                self.errors['delivery_errors'] = supplier_errors
                
        return supplier_errors
        
            
    def validate_delivery_total_quantity(self, total_delivery_quantity, total_allocation_quantity):
        error = None
        # if total_delivery_quantity != total_allocation_quantity:
        #     error = 'Total delivery quantity and allocation quantity mismatch.'
        return error

    def refresh_supplier_pos(self, po_club):
        if po_club.supplierpo_set.filter(state=SupplierPO.DRAFT_STATE).exists():
            NewGenerateSupplierPOs(po_club).create_supplier_pos()

    def get_supplier_delivery_date(self, delivery_date, supplier, po_club):
        supplier_delivery_date = get_object_or_none(SupplierDeliveryDate, {'supplier': supplier, 'po_club': po_club,
                                                                           'confirmed_delivery_date': delivery_date})
        if not supplier_delivery_date:
            supplier_delivery_date = SupplierDeliveryDate.objects.create(
                general_po_supplier__supplier=supplier, po_club=po_club, confirmed_delivery_date=delivery_date
            )
        return supplier_delivery_date

    def update_supplier_price(self, suppliers, general_po_material_quntity, is_po_club_general_po, po_type):
        for supplier_data in suppliers:
            supplier_inquiry_detail = None
            supplier = Supplier.objects.get(pk=supplier_data['id'])
            supplier_price_data = supplier_data['supplier_prices']
            
            supplier_price_id = supplier_price_data.get('id', None)
            supplier_inquiry_detail_id = supplier_price_data.get('supplier_inquiry_detail', None)
            lead_time = supplier_price_data.get('lead_time', None)
            order_price = supplier_price_data.get('order_price', None)
            order_price_units = supplier_price_data.get('order_price_units', None)
            discount = supplier_price_data.get('discount', None)
            excess_threshold = supplier_price_data.get('excess_threshold', None)
            transport_method = supplier_price_data.get('transport_method', None)
            if supplier_inquiry_detail_id:
                supplier_inquiry_detail = SupplierInquiryDetail.objects.get(pk=supplier_inquiry_detail_id)
            
            if supplier_price_id:
                supplier_price = GeneralPOSupplierMaterialPrice.objects.get(pk=supplier_price_id)
                if supplier_inquiry_detail:
                    supplier_price.supplier_inquiry_detail = supplier_inquiry_detail
            else:
                general_po_supplier, created = GeneralPOSupplier.objects.get_or_create(
                    supplier=supplier, general_po=general_po_material_quntity.general_po
                )
                new_supplier_material = general_po_material_quntity.material.supplier_inquiry_material_code.get_related_supplier_material_for_different_supplier_material(general_po_material_quntity.material, True)
                supplier_price = GeneralPOSupplierMaterialPrice.objects.create(
                    general_po_supplier=general_po_supplier,
                    supplier_inquiry_detail=supplier_inquiry_detail,
                    supplier_material=new_supplier_material
                )
            supplier_price.costing_price = order_price #supplier_inquiry_detail.cost_per_unit
            supplier_price.costing_price_units = order_price_units #supplier_inquiry_detail.costing_unit
            supplier_price.lead_time = lead_time
            supplier_price.order_price = order_price
            supplier_price.order_price_units = order_price_units
            supplier_price.discount = discount
            supplier_price.excess_threshold = excess_threshold
            supplier_price.transport_method = transport_method
            supplier_price.general_po_supplier.po_type = po_type
            supplier_price.general_po_supplier.save()
            supplier_price.save()
            self.create_po_allocation(general_po_material_quntity, supplier_data, supplier, supplier_price, is_po_club_general_po)
        return True

    def create_po_allocation(self, general_po_material_quntity, supplier_data, supplier, supplier_price, is_po_club_general_po):
        normalize_material_quantity_unit = general_po_material_quntity.material.material_normalized_measuring_unit
        for delivery in supplier_data['deliveries']:
            delivery_quantity_id = delivery['id']
            confirmed_delivery_date = delivery['confirmed_delivery_date']
            quantity = delivery['quantity']
            transport_method = delivery['transport_method']
            port = delivery.get('port', None)
            po_allocation_quantity_data = delivery['supplierdeliverydatequantitypoallocation_set']
            supplier_delivery_date, created = SupplierDeliveryDate.objects.get_or_create(
                general_po_supplier=supplier_price.general_po_supplier,
                confirmed_delivery_date=confirmed_delivery_date,
                transport_method=transport_method,
                supplier_port_id=port
            )
            if delivery_quantity_id:
                delivery_quantity = SupplierDeliveryDateQuantity.objects.get(pk=delivery_quantity_id)
                delivery_quantity.material_supplier = supplier_price
                delivery_quantity.supplier_delivery_date = supplier_delivery_date
            else:
                delivery_quantity = SupplierDeliveryDateQuantity.objects.create(
                    general_po_material_quantity=general_po_material_quntity,
                    supplier_delivery_date=supplier_delivery_date,
                    material_supplier=supplier_price
                )
            delivery_quantity.quantity = quantity['quantity']
            delivery_quantity.quantity_units = normalize_material_quantity_unit
            delivery_quantity.proforma_invoice_quantity = quantity['quantity']
            delivery_quantity.proforma_invoice_quantity_units = normalize_material_quantity_unit
            delivery_quantity.transport_quantity = quantity['quantity']
            delivery_quantity.transport_quantity_units = normalize_material_quantity_unit
            delivery_quantity.ex_mill_date = delivery.get('exmill_date', None)
            delivery_quantity.save()

            if is_po_club_general_po:
                for row in po_allocation_quantity_data:
                    purchase_order = row['purchase_order']
                    quantity = row['quantity']
                    quantity_units = normalize_material_quantity_unit
                    po_allocation, created = SupplierDeliveryDateQuantityPOAllocation.objects.get_or_create(
                        purchase_order_id=purchase_order,
                        supplier_delivery_date_quantity=delivery_quantity
                    )
                    po_allocation.quantity = quantity
                    po_allocation.quantity_units = quantity_units
                    po_allocation.proforma_invoice_quantity = quantity
                    po_allocation.proforma_invoice_quantity_units = quantity_units
                    po_allocation.save()
        return True

    def update_purchase_order_boms(self, is_po_club_general_po, general_po_material_quntity, po_breakdown_data, material_total_order_quantity):
        if is_po_club_general_po:
            po_bom_ids = []
            for row in po_breakdown_data['purchase_order_boms']:
                    po_bom_id = row['id']
                    order_quantity = row['order_quantity']
                    measuring_unit = row['measuring_unit']
                    po_bom = get_object_or_404(PurchaseOrderBom, pk=po_bom_id)
                    po_bom.order_quantity = order_quantity
                    po_bom.measuring_unit = measuring_unit
                    po_bom.save()
                    po_bom_ids.append(po_bom.id)

            po_boms = PurchaseOrderBom.objects.filter(id__in=po_bom_ids)
            from shared.utils import calculate_queryset_total_normalized_quantity
            normalized_unit = general_po_material_quntity.material.material_normalized_measuring_unit
            total_order_quantity = calculate_queryset_total_normalized_quantity(po_boms, normalized_unit, 'order_quantity', 'measuring_unit')
            general_po_material_quntity.order_quantity = total_order_quantity
            general_po_material_quntity.save()
        else:
            general_po_material_quntity.order_quantity = material_total_order_quantity
            general_po_material_quntity.save()

    def validate_data(self, po_breakdown_data, suppliers, material_total_order_quantity, is_po_club_general_po, order_quantity_discrepancy_reason, order_quantity_discrepancy_other, po_type):
        if is_po_club_general_po:
            self.validate_po_bom_order_quantity(po_breakdown_data, suppliers)
        else:
            self.validate_total_order_quantity(suppliers, material_total_order_quantity)
        self.validate_supplier_price(suppliers)
        self.validate_excess_threshold(suppliers)
        self.validate_delivery_errors(suppliers, is_po_club_general_po)
        self.validate_order_quantity_discrepancy_other(order_quantity_discrepancy_reason, order_quantity_discrepancy_other)

    def validate_order_quantity_discrepancy_other(self, order_quantity_discrepancy_reason, order_quantity_discrepancy_other):
        if order_quantity_discrepancy_reason == GeneralPOMaterialQuantity.OTHER_DISCREPANCY_CHOICE and not order_quantity_discrepancy_other:
            self.errors['order_quantity_discrepancy_other'] = "Other Discrepancy reason cannot be empty."

    def post(self, request, *args, **kwargs):
        general_po_material_quntity = get_object_or_404(GeneralPOMaterialQuantity, pk=self.kwargs.get('general_po_material_quantity_id'))
        is_po_club_general_po = general_po_material_quntity.general_po.is_po_club_general_po()
        po_breakdown_data = request.data.get('total_value_with_po_breakdown', {})
        suppliers = request.data.get('suppliers', [])
        material_total_order_quantity = request.data.get('material_total_order_quantity', None)

        material_total_order_quantity = request.data.get('material_total_order_quantity', None)
        order_quantity_discrepancy_reason = request.data.get('order_quantity_discrepancy_reason', None)
        order_quantity_discrepancy_other = request.data.get('order_quantity_discrepancy_other', None)
        completed = request.data.get('completed', False)
        po_type = request.data.get('po_type', False)

        self.validate_data(po_breakdown_data, suppliers, material_total_order_quantity, is_po_club_general_po, order_quantity_discrepancy_reason, order_quantity_discrepancy_other, po_type)

        if self.errors:
            general_po_material_quntity.completed = False
            general_po_material_quntity.save()
            http_response = Response(self.errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            general_po_material_quntity.order_quantity_discrepancy_reason = order_quantity_discrepancy_reason
            if order_quantity_discrepancy_reason == GeneralPOMaterialQuantity.OTHER_DISCREPANCY_CHOICE:
                general_po_material_quntity.order_quantity_discrepancy_other = order_quantity_discrepancy_other
            else:
                general_po_material_quntity.order_quantity_discrepancy_other = None
            general_po_material_quntity.completed = completed
            general_po_material_quntity.save()

            self.update_purchase_order_boms(is_po_club_general_po, general_po_material_quntity, po_breakdown_data, material_total_order_quantity)
            self.update_supplier_price(suppliers, general_po_material_quntity, is_po_club_general_po, po_type)
            http_response = Response({
                'success': True,
                'general_po_material_quantity_id': general_po_material_quntity.id,
                'material_id': general_po_material_quntity.material.id,
                'quantity': general_po_material_quntity.order_quantity,
                'quantity_units_display': general_po_material_quntity.get_order_quantity_units_display(),
                'completed': general_po_material_quntity.completed,
            })
        return http_response
    

class GeneralPOSupplierMaterialPriceDeleteView(generics.RetrieveDestroyAPIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]
    serializer_class = GeneralPOSupplierMaterialPriceSerializer
    queryset = GeneralPOSupplierMaterialPrice.objects.all()


class SupplierDeliveryDateQuantityPOAllocationDeleteView(generics.RetrieveDestroyAPIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]
    serializer_class = SupplierDeliveryDateQuantityPOAllocationSerializer
    queryset = SupplierDeliveryDateQuantityPOAllocation.objects.all()


class GeneralPOMaterialDetail(APIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]

    def get_supplier_prices(self, general_po_material_quantity, material):
        material_price_id = general_po_material_quantity.default_material_supplier.id
        supplier_prices = GeneralPOSupplierMaterialPrice.objects.filter(
            id=material_price_id,
            supplier_material__customer_brand_material=material
        ).distinct().order_by('supplier_inquiry_detail__supplier_inquiry__supplier', 'supplier_inquiry_detail__cutting_width')

        # if not supplier_prices:
        #     supplier_prices = GeneralPOSupplierMaterialPrice.objects.filter(
        #         general_po_supplier__general_po=general_po,
        #         supplier_material__customer_brand_material__material_code=material.material_code
        #     ).order_by('supplier_inquiry_detail__supplier_inquiry__supplier', 'supplier_inquiry_detail__cutting_width')
        return supplier_prices
    
    def get_suppliers(self, general_po_material_quantity, material):
        supplier_ids = SupplierDeliveryDateQuantity.objects.filter(
            general_po_material_quantity__material=material
        ).values_list('material_supplier__supplier_inquiry_detail__supplier_inquiry__supplier', flat=True)
        qs = Supplier.objects.filter(id__in=supplier_ids)
        return qs
    
    def get_total_value_with_po_breakdown(self, general_po_material_quantity):
        material = general_po_material_quantity.material
        purchase_orders = general_po_material_quantity.general_po.po_club.get_purchase_orders()
        purchase_order_boms = PurchaseOrderBom.objects.filter(purchase_order__in=purchase_orders, material=material).order_by('-id')
        total_value_with_po_breakdown = {
            'id': general_po_material_quantity.id,
            'quantity': general_po_material_quantity.quantity,
            'measuring_unit': general_po_material_quantity.measuring_unit,
            'measuring_unit_display': general_po_material_quantity.get_measuring_unit_display(),
            'order_quantity': general_po_material_quantity.order_quantity,
            'order_quantity_units': general_po_material_quantity.order_quantity_units,
            'order_quantity_units_display': general_po_material_quantity.get_order_quantity_units_display(),
            'purchase_order_boms': PurchaseOrderBomBasicSerializer(purchase_order_boms, many=True).data,
        }
        return total_value_with_po_breakdown
    
    def get_is_default_supplier(self, general_po_material_quantity, supplier_price):
        return general_po_material_quantity.default_material_supplier == supplier_price
    
    def get_purchase_orders(self, general_po_material_quantity):
        data = []
        if general_po_material_quantity.general_po.is_po_club_general_po():
            purchase_orders = general_po_material_quantity.general_po.po_club.get_purchase_orders()
            for purchase_order in purchase_orders:
                data.append({
                    'id': purchase_order.id,
                    'name': purchase_order.display_number
                })
        return data

    def get_delivery_breakdown(self, general_po_material_quantity):
        from shared.utils import calculate_queryset_total_normalized_quantity, get_quantity_dictionary
        data = []
        supplier_prices = GeneralPOSupplierMaterialPrice.objects.filter(
            general_po_supplier__general_po=general_po_material_quantity.general_po,
            supplier_material__customer_brand_material=general_po_material_quantity.material
        )
        for supplier_price in supplier_prices:
            supplier_data = {
                'id': supplier_price.general_po_supplier.supplier.id,
                'supplier_name': supplier_price.general_po_supplier.supplier.name,
                'order_price': supplier_price.order_price,
                'order_price_units': supplier_price.get_order_price_units_display(),
                'total_quantity': get_quantity_dictionary(calculate_queryset_total_normalized_quantity(supplier_price.supplierdeliverydatequantity_set.all(), 
                                                                                                       supplier_price.supplier_material.customer_brand_material.material_normalized_measuring_unit, 'quantity', 'quantity_units'), 'quantity_units'),
                'width': supplier_price.cutting_width,
                'width_display_units': supplier_price.get_cutting_width_unit_display(),
                'transport_method': supplier_price.transport_method,
                'transport_type_display': supplier_price.get_transport_method_display(),
                'cost_per_unit_type': supplier_price.get_cost_per_unit_type_display(),
                'deliveries': []
            }
            quantites = supplier_price.supplierdeliverydatequantity_set.all()
            for quantity in quantites:
                quantity_data = {
                    'id': quantity.id,
                    'quantity': quantity.get_normalized_quantity(),
                    'delivery_date': quantity.supplier_delivery_date.confirmed_delivery_date if quantity.supplier_delivery_date else None,
                    'exmill_date': quantity.ex_mill_date,
                    'transport_method': quantity.supplier_delivery_date.transport_method if quantity.supplier_delivery_date else None,
                    'po_allocations': []
                }
                po_allocations = quantity.supplierdeliverydatequantitypoallocation_set.all()
                for po_allocation in po_allocations:
                    po_allocation_data = {
                        'id': po_allocation.id,
                        'purchase_order_id': po_allocation.purchase_order.id,
                        'purchase_order_display_number': po_allocation.purchase_order.display_number,
                        'quantity': po_allocation.get_normalized_quantity(),
                    }
                    quantity_data['po_allocations'].append(po_allocation_data)
                supplier_data['deliveries'].append(quantity_data)
            data.append(supplier_data)
        return data
    
    def get(self, request, *args, **kwargs):
        response = {}
        general_po_material_quantity_id = self.kwargs.get('general_po_material_quantity_id')
        general_po_material_quantity = get_object_or_404(GeneralPOMaterialQuantity, pk=general_po_material_quantity_id)
        print(general_po_material_quantity)
        is_po_club_general_po = general_po_material_quantity.general_po.is_po_club_general_po()
        material = general_po_material_quantity.material
        response = {
            'material': CustomerBrandMaterialBasicSerializer(material).data,
            'total_value_with_po_breakdown': self.get_total_value_with_po_breakdown(general_po_material_quantity) if is_po_club_general_po else None,
            'material_total_quantity': general_po_material_quantity.quantity if not is_po_club_general_po else None,
            'material_total_quantity_units': general_po_material_quantity.get_measuring_unit_display() if not is_po_club_general_po else None,
            'material_total_order_quantity': general_po_material_quantity.order_quantity if not is_po_club_general_po else None,
            'material_total_order_quantity_units': general_po_material_quantity.get_order_quantity_units_display() if not is_po_club_general_po else None,
            'suppliers': [],
            'po_list': self.get_purchase_orders(general_po_material_quantity),
            'delivery_breakdown': self.get_delivery_breakdown(general_po_material_quantity), #TODO Ask Dasith sir about commenting this line
            'order_quantity_discrepancy_reason': general_po_material_quantity.order_quantity_discrepancy_reason,
            'order_quantity_discrepancy_other': general_po_material_quantity.order_quantity_discrepancy_other,
            'completed': general_po_material_quantity.completed,
            'po_club_id': general_po_material_quantity.general_po.po_club.id if general_po_material_quantity.general_po.po_club else None,
            'po_type': general_po_material_quantity.default_material_supplier.general_po_supplier.po_type,
        }

        supplier_prices = self.get_supplier_prices(general_po_material_quantity, material) 
        
        for supplier_price in supplier_prices:
            deliveries = supplier_price.get_deliveries().filter(general_po_material_quantity=general_po_material_quantity)
            response['suppliers'].append({
                'id': supplier_price.supplier_material.supplier.id,
                'general_po_material_price_id': supplier_price.id,
                'general_po_id': supplier_price.general_po_supplier.general_po.id,
                'supplier_name': supplier_price.supplier_material.supplier.name,
                'supplier_material_reference_code': supplier_price.supplier_material.supplier_material_reference_code,
                'cutting_width': supplier_price.cutting_width,
                'cutting_width_unit_display': supplier_price.get_cutting_width_unit_display(),
                'no_of_deliveries': supplier_price.get_delivery_count(),
                'delivery_type': supplier_price.get_delivery_type(),
                'is_default_supplier': self.get_is_default_supplier(general_po_material_quantity, supplier_price),
                'supplier_prices': GeneralPOSupplierMaterialPriceSerializer(supplier_price, many=False).data,
                'ship_mode': supplier_price.get_ship_mode_display(),
                'cost_per_unit_type': supplier_price.get_cost_per_unit_type_display(),
                'deliveries': SupplierDeliveryDateQuantitySerializer(deliveries, many=True).data, # THIS IS WHERE THE BUG IS. Why is it not being filtered here?
            })

        return Response(response)
    

class GeneralPOMaterialList(APIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]

    def get_supplier_prices(self, general_po, material):
        supplier_prices = GeneralPOSupplierMaterialPrice.objects.filter(
            general_po_supplier__general_po=general_po,
            supplier_material__customer_brand_material=material
        ).distinct().order_by('supplier_inquiry_detail__supplier_inquiry__supplier', 'supplier_inquiry_detail__cutting_width')

        if not supplier_prices:
            supplier_prices = GeneralPOSupplierMaterialPrice.objects.filter(
                general_po_supplier__general_po=general_po,
                supplier_material__customer_brand_material__material_code=material.material_code
            ).order_by('supplier_inquiry_detail__supplier_inquiry__supplier', 'supplier_inquiry_detail__cutting_width')
        return supplier_prices
    
    def get_suppliers(self, general_po_material_quantity, material):
        supplier_ids = SupplierDeliveryDateQuantity.objects.filter(
            general_po_material_quantity__material=material
        ).values_list('material_supplier__supplier_inquiry_detail__supplier_inquiry__supplier', flat=True)
        qs = Supplier.objects.filter(id__in=supplier_ids)
        return qs
    
    def get_total_value_with_po_breakdown(self, general_po_material_quantity):
        material = general_po_material_quantity.material
        purchase_orders = general_po_material_quantity.general_po.po_club.get_purchase_orders()
        purchase_order_boms = PurchaseOrderBom.objects.filter(purchase_order__in=purchase_orders, material=material).order_by('-id')
        total_value_with_po_breakdown = {
            'id': general_po_material_quantity.id,
            'quantity': general_po_material_quantity.quantity,
            'measuring_unit': general_po_material_quantity.measuring_unit,
            'measuring_unit_display': general_po_material_quantity.get_measuring_unit_display(),
            'order_quantity': general_po_material_quantity.order_quantity,
            'order_quantity_units': general_po_material_quantity.order_quantity_units,
            'order_quantity_units_display': general_po_material_quantity.get_order_quantity_units_display(),
            'purchase_order_boms': PurchaseOrderBomBasicSerializer(purchase_order_boms, many=True).data,
        }
        return total_value_with_po_breakdown
    
    def get_is_default_supplier(self, general_po_material_quantity, supplier_price):
        return general_po_material_quantity.default_material_supplier == supplier_price
    
    def get_purchase_orders(self, general_po_material_quantity):
        data = []
        if general_po_material_quantity.general_po.is_po_club_general_po():
            purchase_orders = general_po_material_quantity.general_po.po_club.get_purchase_orders()
            for purchase_order in purchase_orders:
                data.append({
                    'id': purchase_order.id,
                    'name': purchase_order.display_number
                })
        return data

    def get_delivery_breakdown(self, general_po_material_quantity):
        from shared.utils import calculate_queryset_total_normalized_quantity, get_quantity_dictionary
        data = []
        supplier_prices = GeneralPOSupplierMaterialPrice.objects.filter(
            general_po_supplier__general_po=general_po_material_quantity.general_po,
            supplier_material__customer_brand_material=general_po_material_quantity.material
        )
        for supplier_price in supplier_prices:
            supplier_data = {
                'id': supplier_price.general_po_supplier.supplier.id,
                'supplier_name': supplier_price.general_po_supplier.supplier.name,
                'order_price': supplier_price.order_price,
                'order_price_units': supplier_price.get_order_price_units_display(),
                'total_quantity': get_quantity_dictionary(calculate_queryset_total_normalized_quantity(supplier_price.supplierdeliverydatequantity_set.all(), 
                                                                                                       supplier_price.supplier_material.customer_brand_material.material_normalized_measuring_unit, 'quantity', 'quantity_units'), 'quantity_units'),
                'width': supplier_price.supplier_inquiry_detail.cutting_width,
                'width_display_units': supplier_price.supplier_inquiry_detail.get_cutting_width_unit_display(),
                'transport_method': supplier_price.transport_method,
                'transport_type_display': supplier_price.get_transport_method_display(),
                'cost_per_unit_type': supplier_price.supplier_inquiry_detail.get_cost_per_unit_type_display(),
                'deliveries': []
            }
            quantites = supplier_price.supplierdeliverydatequantity_set.all()
            for quantity in quantites:
                quantity_data = {
                    'id': quantity.id,
                    'quantity': quantity.get_normalized_quantity(),
                    'delivery_date': quantity.supplier_delivery_date.confirmed_delivery_date if quantity.supplier_delivery_date else None,
                    'exmill_date': quantity.ex_mill_date,
                    'transport_method': quantity.supplier_delivery_date.transport_method if quantity.supplier_delivery_date else None,
                    'po_allocations': []
                }
                po_allocations = quantity.supplierdeliverydatequantitypoallocation_set.all()
                for po_allocation in po_allocations:
                    po_allocation_data = {
                        'id': po_allocation.id,
                        'purchase_order_id': po_allocation.purchase_order.id,
                        'purchase_order_display_number': po_allocation.purchase_order.display_number,
                        'quantity': po_allocation.get_normalized_quantity(),
                    }
                    quantity_data['po_allocations'].append(po_allocation_data)
                supplier_data['deliveries'].append(quantity_data)
            data.append(supplier_data)
        return data
    
    def post(self, request, *args, **kwargs):
        data = []
        general_po_material_quantity_ids = request.data.get('general_po_material_quantity_ids', [])
        general_po_material_quantities = GeneralPOMaterialQuantity.objects.filter(id__in=general_po_material_quantity_ids)
        for general_po_material_quantity in general_po_material_quantities:
            response = {}
            is_po_club_general_po = general_po_material_quantity.general_po.is_po_club_general_po()
            material = general_po_material_quantity.material
            response = {
                'material': CustomerBrandMaterialBasicSerializer(material).data,
                'total_value_with_po_breakdown': self.get_total_value_with_po_breakdown(general_po_material_quantity) if is_po_club_general_po else None,
                'material_total_quantity': general_po_material_quantity.quantity if not is_po_club_general_po else None,
                'material_total_quantity_units': general_po_material_quantity.get_measuring_unit_display() if not is_po_club_general_po else None,
                'material_total_order_quantity': general_po_material_quantity.order_quantity if not is_po_club_general_po else None,
                'material_total_order_quantity_units': general_po_material_quantity.get_order_quantity_units_display() if not is_po_club_general_po else None,
                'suppliers': [],
                'po_list': self.get_purchase_orders(general_po_material_quantity),
                'delivery_breakdown': self.get_delivery_breakdown(general_po_material_quantity), #TODO Ask Dasith sir about commenting this line
                'order_quantity_discrepancy_reason': general_po_material_quantity.order_quantity_discrepancy_reason,
                'order_quantity_discrepancy_other': general_po_material_quantity.order_quantity_discrepancy_other,
                'completed': general_po_material_quantity.completed,
                'po_club_id': general_po_material_quantity.general_po.po_club.id if general_po_material_quantity.general_po.po_club else None,
                'po_type': general_po_material_quantity.default_material_supplier.general_po_supplier.po_type,
            }

            supplier_prices = self.get_supplier_prices(general_po_material_quantity.general_po, material) 
            
            for supplier_price in supplier_prices:
                deliveries = supplier_price.get_deliveries().filter(general_po_material_quantity=general_po_material_quantity)
                response['suppliers'].append({
                    'id': supplier_price.supplier_inquiry_detail.supplier_inquiry.supplier.id,
                    'general_po_material_price_id': supplier_price.id,
                    'general_po_id': supplier_price.general_po_supplier.general_po.id,
                    'supplier_name': supplier_price.supplier_inquiry_detail.supplier_inquiry.supplier.name,
                    'supplier_material_reference_code': supplier_price.supplier_inquiry_detail.supplier_inquiry_material_code.supplier_material_reference_code,
                    'cutting_width': supplier_price.supplier_inquiry_detail.cutting_width,
                    'cutting_width_unit_display': supplier_price.supplier_inquiry_detail.get_cutting_width_unit_display(),
                    'no_of_deliveries': supplier_price.get_delivery_count(),
                    'delivery_type': supplier_price.get_delivery_type(),
                    'is_default_supplier': self.get_is_default_supplier(general_po_material_quantity, supplier_price),
                    'supplier_prices': GeneralPOSupplierMaterialPriceSerializer(supplier_price, many=False).data,
                    'ship_mode': supplier_price.supplier_inquiry_detail.get_ship_mode_display(),
                    'cost_per_unit_type': supplier_price.supplier_inquiry_detail.get_cost_per_unit_type_display(),
                    'deliveries': SupplierDeliveryDateQuantityListSerializer(deliveries, many=True).data, # THIS IS WHERE THE BUG IS. Why is it not being filtered here?
                })
            data.append(response)
        return Response(data)
    

class GeneralPOMaterialListSaveView(APIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]

    def update_supplier_price(self, suppliers, general_po_material_quntity):
        for supplier_data in suppliers:
            supplier = Supplier.objects.get(pk=supplier_data['id'])
            supplier_price_data = supplier_data['supplier_prices']
            
            supplier_price_id = supplier_price_data.get('id', None)
            supplier_inquiry_detail_id = supplier_price_data.get('supplier_inquiry_detail', None)
            lead_time = supplier_price_data.get('lead_time', None)
            order_price = supplier_price_data.get('order_price', None)
            order_price_units = supplier_price_data.get('order_price_units', None)
            discount = supplier_price_data.get('discount', None)
            excess_threshold = supplier_price_data.get('excess_threshold', None)
            transport_method = supplier_price_data.get('transport_method', None)
            supplier_inquiry_detail = SupplierInquiryDetail.objects.get(pk=supplier_inquiry_detail_id)
            
            supplier_price = GeneralPOSupplierMaterialPrice.objects.get(pk=supplier_price_id)
            supplier_price.supplier_inquiry_detail = supplier_inquiry_detail
            
            supplier_price.costing_price = supplier_inquiry_detail.cost_per_unit
            supplier_price.costing_price_units = supplier_inquiry_detail.costing_unit
            supplier_price.lead_time = lead_time
            supplier_price.order_price = order_price
            supplier_price.order_price_units = order_price_units
            supplier_price.discount = discount
            supplier_price.excess_threshold = excess_threshold
            supplier_price.transport_method = transport_method
            supplier_price.general_po_supplier.save()
            supplier_price.save()
            self.create_po_allocation(general_po_material_quntity, supplier_data, supplier, supplier_price)
        return True

    def create_po_allocation(self, general_po_material_quntity, supplier_data, supplier, supplier_price,):
        normalize_material_quantity_unit = general_po_material_quntity.material.material_normalized_measuring_unit
        for delivery in supplier_data['deliveries']:
            delivery_quantity_id = delivery['id']
            confirmed_delivery_date = delivery['confirmed_delivery_date']
            quantity = delivery['quantity']
            transport_method = delivery['transport_method']
            port = delivery.get('port', None)
            po_allocation_quantity_data = delivery['po_breakdown']
            supplier_delivery_date, created = SupplierDeliveryDate.objects.get_or_create(
                general_po_supplier=supplier_price.general_po_supplier,
                confirmed_delivery_date=confirmed_delivery_date,
                transport_method=transport_method,
                supplier_port_id=port
            )
            if delivery_quantity_id:
                delivery_quantity = SupplierDeliveryDateQuantity.objects.get(pk=delivery_quantity_id)
                delivery_quantity.material_supplier = supplier_price
                delivery_quantity.supplier_delivery_date = supplier_delivery_date
            else:
                delivery_quantity = SupplierDeliveryDateQuantity.objects.create(
                    general_po_material_quantity=general_po_material_quntity,
                    supplier_delivery_date=supplier_delivery_date,
                    material_supplier=supplier_price
                )
            delivery_quantity.quantity = quantity['quantity']
            delivery_quantity.quantity_units = normalize_material_quantity_unit
            delivery_quantity.proforma_invoice_quantity = quantity['quantity']
            delivery_quantity.proforma_invoice_quantity_units = normalize_material_quantity_unit
            delivery_quantity.transport_quantity = quantity['quantity']
            delivery_quantity.transport_quantity_units = normalize_material_quantity_unit
            delivery_quantity.ex_mill_date = delivery.get('exmill_date', None)
            delivery_quantity.save()

            for row in po_allocation_quantity_data:
                purchase_order = row['purchase_order']
                quantity = row['quantity']
                quantity_units = normalize_material_quantity_unit
                po_allocation, created = SupplierDeliveryDateQuantityPOAllocation.objects.get_or_create(
                    purchase_order_id=purchase_order,
                    supplier_delivery_date_quantity=delivery_quantity
                )
                po_allocation.quantity = quantity
                po_allocation.quantity_units = quantity_units
                po_allocation.proforma_invoice_quantity = quantity
                po_allocation.proforma_invoice_quantity_units = quantity_units
                po_allocation.save()

            self.update_purchase_order_boms(po_allocation_quantity_data, general_po_material_quntity)
        return True

    def update_purchase_order_boms(self, po_breakdown_data, general_po_material_quntity):
        po_bom_ids = []
        for row in po_breakdown_data:
                po_bom_id = row['id']
                order_quantity = row['quantity']
                measuring_unit = row['measuring_unit']
                po_bom = get_object_or_404(PurchaseOrderBom, pk=po_bom_id)
                po_bom.order_quantity = order_quantity
                po_bom.measuring_unit = measuring_unit
                po_bom.save()
                po_bom_ids.append(po_bom.id)

        po_boms = PurchaseOrderBom.objects.filter(id__in=po_bom_ids)
        from shared.utils import calculate_queryset_total_normalized_quantity
        normalized_unit = general_po_material_quntity.material.material_normalized_measuring_unit
        total_order_quantity = calculate_queryset_total_normalized_quantity(po_boms, normalized_unit, 'order_quantity', 'measuring_unit')
        general_po_material_quntity.order_quantity = total_order_quantity
        general_po_material_quntity.save()

    def validate_data(self, data):
        errors = {}
        has_errors = False
        main_index = 0
        for row in data:
            suppliers = row.get('suppliers', [])
            supplier_errors = {}
            index = 0
            for supplier in suppliers:
                material_key = str(index)
                
                delivery_errors = {}
                for delivery in supplier.get("deliveries", []):
                    if not delivery.get("transport_method"):
                        if delivery['id'] not in delivery_errors:
                            delivery_errors[delivery['id']] = {}
                        delivery_errors[delivery['id']]["transport_method"] = "Select transport method"
                        has_errors = True

                    if not delivery.get("confirmed_delivery_date"):
                        if delivery['id'] not in delivery_errors:
                            delivery_errors[delivery['id']] = {}
                        delivery_errors[delivery['id']]["confirmed_delivery_date"] = "Select delivery date"
                        has_errors = True

                    if not delivery.get("exmill_date"):
                        if delivery['id'] not in delivery_errors:
                            delivery_errors[delivery['id']] = {}
                        delivery_errors[delivery['id']]["exmill_date"] = "Select exmill date"
                        has_errors = True

                    if not delivery.get("port"):
                        if delivery['id'] not in delivery_errors:
                            delivery_errors[delivery['id']] = {}
                        delivery_errors[delivery['id']]["port"] = "Select port"
                        has_errors = True

                    po_errors = {}
                    for po in delivery.get("po_breakdown", []):
                        if not po.get("quantity"):
                            po_errors[str(po.get("id"))] = {
                                "quantity_error": "Enter quantity"
                            }
                            has_errors = True

                    if po_errors:
                        if delivery['id'] not in delivery_errors:
                            delivery_errors[delivery['id']] = {}
                        delivery_errors[delivery['id']]["po_errors"] = po_errors

                    if delivery_errors:
                        if material_key not in supplier_errors:
                            supplier_errors[material_key] = {}
                        supplier_errors[material_key]["delivery_errors"] = delivery_errors

                index += 1
                if supplier_errors:
                    if main_index not in errors:
                        errors[main_index] = {}
                    errors[main_index] = supplier_errors
            main_index += 1
        return has_errors, errors

    def post(self, request, *args, **kwargs):
        has_errors, errors = self.validate_data(request.data.get('materials'))
        if has_errors:
            http_response = Response(errors, status=status.HTTP_400_BAD_REQUEST)
        else:
            for row in request.data.get('materials'):
                po_breakdown_data = row.get('total_value_with_po_breakdown', {})
                general_po_material_quntity = get_object_or_404(GeneralPOMaterialQuantity, pk=po_breakdown_data.get('id'))
                suppliers = row.get('suppliers', [])
                self.update_supplier_price(suppliers, general_po_material_quntity)
                general_po_material_quntity.completed = True
                general_po_material_quntity.save()
            http_response = Response({'success': True})            
        return http_response
    

class GeneralPOMaterialDeliveryMetaDataView(APIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]
    
    def get(self, request, po_club_id, supplier_id, material_id):
        response = {}
        po_club = get_object_or_404(ActualPOClub, pk=po_club_id)
        supplier = get_object_or_404(Supplier, pk=supplier_id)
        material = get_object_or_404(CustomerBrandMaterial, pk=material_id)

        supplier_delivery_date_quantities = SupplierDeliveryDateQuantity.objects.filter(
            material_supplier__general_po_supplier__supplier=supplier, general_po_material_quantity__material=material,
            general_po_material_quantity__general_po__po_club=po_club
        )
        if supplier_delivery_date_quantities:
            supplier_delivery_date_quantity = supplier_delivery_date_quantities[0]
            response = {
                'ex_mill_date': supplier_delivery_date_quantity.ex_mill_date,
                'delivery_date': supplier_delivery_date_quantity.supplier_delivery_date.confirmed_delivery_date if supplier_delivery_date_quantity.supplier_delivery_date else None,
                'ship_mode': supplier_delivery_date_quantity.material_supplier.supplier_inquiry_detail.ship_mode if supplier_delivery_date_quantity.material_supplier.supplier_inquiry_detail else None,
                'port': supplier_delivery_date_quantity.supplier_delivery_date.supplier_port_id if supplier_delivery_date_quantity.supplier_delivery_date else None,
            }
        return Response(response)


class GeneralPOPOClubMaterialQuantityListView(generics.ListAPIView, OrderMixin):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]
    serializer_class = GeneralPOMaterialQuantityListSerializer

    def get_queryset(self):
        club_id = self.kwargs.get('club_id')
        po_club = self.get_po_club_or_raise_http404(club_id)
        state = self.request.query_params.get('state', None)

        qs = GeneralPOMaterialQuantity.objects.filter(general_po__po_club=po_club).order_by(
            'material__material_detail__generic_material__user_material__display_order', 'material__id')
        if state == 'supplier_po_finalized':
            qs = qs.filter(supplierpogeneralpomaterialquantity__isnull=False)
        if state == 'supplier_po_pending':
            qs = qs.filter(supplierpogeneralpomaterialquantity__isnull=True) #True
        return qs
    

class GeneralPOPOClubMaterialQuantityExcelListView(APIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]

    def calcuate_cost(self, material, cost_per_unit, unit, costing_consumption_ratio, wastage):
        total_cost = 0.00
        normalized_cost_per_unit = convert_per_unit_cost(material.material_normalized_measuring_unit, cost_per_unit, unit)

        total_consumption = costing_consumption_ratio + (costing_consumption_ratio * wastage / 100)
        total_cost = normalized_cost_per_unit['cost'] * float(total_consumption)
        total_cost = round(total_cost, 5)

        return total_cost

    def create_bom_excel(self, po_club, data, purchase_order_quantity_data, po_club_quantity_data):
        from django.http import HttpResponse
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.drawing.image import Image
        from datetime import datetime
        from io import BytesIO

        logo_path="shared/static/images/ritz/RITZ-LOGO.jpg"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bill of Materials"

        # Styles
        header_font = Font(bold=True, size=12)
        subheader_font = Font(bold=True, size=11)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
        header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        try:
            img = Image(logo_path)
            img.width = 100
            img.height = 100
            ws.merge_cells(f'A{1}:A{5}')
            ws['A1'].alignment = center_align
            ws.add_image(img, 'A1')
        except FileNotFoundError:
            print(f"Warning: Logo file not found at {logo_path}. Proceeding without logo.")

        start_row = 6
        header_start_row = 1

        ws.merge_cells(f'B{header_start_row}:N{header_start_row + 4}')
        ws[f'B{header_start_row}'] = (
                        "Ritz Design Center (Pvt) Ltd.\n"
                        "Bill of Materials\n"
                        "Approved BOM\n"
                        "Tel: +94 000000000\n"
        )
        ws[f'B{header_start_row}'].font = header_font
        ws[f'B{header_start_row}'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        general_info = [
            ("PO Club", data.get('po_club_number', '')),
            ("Style Number", data.get('style')),
            ("Customer", data.get('customer', '')),
            ("Order", data.get('order')),
            ("Costing", data.get('costing', '')),
        ]

        row = start_row
        for label, value in general_info:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = subheader_font
            ws[f'A{row}'].alignment = left_align
            ws.merge_cells(f'B{row}:L{row}')
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = subheader_font
            ws[f'B{row}'].alignment = left_align
            row += 1

        # Purchase Order Quantities

        row += 2
        ws[f'A{row}'] = "System PO Number"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].alignment = left_align

        ws[f'B{row}'] = "Customer PO Number"
        ws[f'B{row}'].font = subheader_font
        ws[f'B{row}'].alignment = left_align

        ws[f'C{row}'] = "Quantity"
        ws[f'C{row}'].font = subheader_font
        ws[f'C{row}'].alignment = left_align

        row += 1

        for po_row in purchase_order_quantity_data:
            ws[f'A{row}'] = po_row['purchase_order']
            ws[f'A{row}'].alignment = left_align

            ws[f'B{row}'] = po_row['customer_po_name']
            ws[f'B{row}'].alignment = left_align

            ws[f'C{row}'] = po_row['quantity']
            ws[f'C{row}'].alignment = left_align
            row += 1
        # END Purchase Order Quantities

        # Start PO Club Quantities
        row += 2

        ws[f'A{row}'] = "Size"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].alignment = left_align

        row += 1
        ws[f'A{row}'] = "Quantity"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].alignment = left_align

        row += 1
        ws[f'A{row}'] = "Total Pack Quantity"
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].alignment = left_align

        col = 2 
        for club_row in po_club_quantity_data:
            ws.cell(row=row-2, column=col,  value=club_row["po_pack_display"]).alignment = left_align
            ws.cell(row=row-2, column=col).font = subheader_font 
            ws.cell(row=row-1, column=col, value=club_row["quantity"]).alignment = left_align
            total_pack_quantity = club_row["quantity"] * po_club.pre_costing.order.quantity_per_pack
            ws.cell(row=row, column=col, value=total_pack_quantity).alignment = left_align
            col += 1

        row += 1
        # END PO Club Quantities

        headers = ["Cost Type", "Item Code", "Item Description", "Supplier", "Size", "Per Unit Qty", "Consumption", "Cost",  "Req Qty", "Wastage (%)", "Qty with Wastage", "GRN Quantity", "Balance To GRN", "Item Value", ]
        
        row += 1
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = subheader_font
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = border

        sections = [
            ("Fabric Cost", data.get('fabric_cost', [])),
            ("Sewing Trim Cost", data.get('sewing_trim_cost', [])),
            ("Packing Trim Cost", data.get('packing_trim_cost', []))
        ]

        for section_name, items in sections:
            row += 1
            ws.merge_cells(f'A{row}:N{row}')
            cell = ws[f'A{row}']
            cell.value = section_name
            cell.font = subheader_font
            cell.fill = header_fill
            cell.alignment = left_align
            cell.border = border

            for item in items:
                row += 1
                row_data = [
                    item.get('material_label', ''),
                    item.get('item_code', ''),
                    item.get('item_description', ''),
                    item.get('supplier', ''),
                    item.get('size', ''),
                    item.get('per_unit_qty', 0),
                    item.get('consumption', 0),
                    item.get('cost', 0),
                    item.get('req_qty', 0),
                    item.get('wastage', 0),
                    item.get('qty_with_wastage', 0),
                    item.get('grn_quantity', 0),
                    item.get('variance', 0),
                    item.get('item_value', 0),
                ]
                for col, value in enumerate(row_data, start=1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.alignment = left_align
                    cell.border = border
                    if col in [5, 6, 7, 8, 9]:
                        cell.number_format = '0.00000'

                    if col in [10, 11, 12]:
                        cell.number_format = '0.00'

            # Section Total
            if items:
                row += 1
                ws.merge_cells(f'A{row}:M{row}')
                cell = ws[f'N{row}']
                cell.value = sum(item.get('item_value', 0) for item in items)
                cell.font = subheader_font
                cell.alignment = right_align
                cell.number_format = '0.00'
                cell.border = border

        # Grand Total
        row += 1
        ws.merge_cells(f'A{row}:M{row}')
        cell = ws[f'N{row}']
        cell.value = sum(sum(item.get('item_value', 0) for item in section) for _, section in sections)
        cell.font = header_font
        cell.alignment = right_align
        cell.number_format = '0.00'
        cell.border = border
        
        for col in range(5, 12):
            ws.column_dimensions[get_column_letter(col)].width = 15
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 70
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['L'].width = 15

        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=purchase_order_report.xlsx'
        return response
    
    def get_po_club_pack_quantity_data(self, po_club):
        aggregated_data = {}
        purchase_orders = po_club.get_purchase_orders()
        total_order_quantity = 0
        for purchase_order in purchase_orders:
            po_packs = purchase_order.get_po_packs().order_by('po_size__order_size__size__sorting_order')
            for po_pack in po_packs:
                key = (
                    po_pack.po_country.order_country.id,
                    po_pack.po_size.order_size.id,
                    po_pack.po_colorway.order_colorway.id
                )

                if key in aggregated_data:
                    aggregated_data[key]['quantity'] += po_pack.quantity
                    total_order_quantity += po_pack.quantity
                else:
                    total_order_quantity += po_pack.quantity
                    aggregated_data[key] = {
                        'po_pack_id': po_pack.id,
                        'po_pack_display': po_pack.po_size.po_size_name,
                        'quantity': po_pack.quantity,
                    }

        return list(aggregated_data.values())
    
    def get_purchase_order_quantity_data(self, po_club):
        from django.db.models import Sum
        data = []
        purchase_orders = po_club.get_purchase_orders()
        for po in purchase_orders:
            quantity = po.get_po_packs().aggregate(quantity=Sum('quantity'))['quantity'] or 0
            data.append({
                'purchase_order': po.display_number,
                'quantity': quantity,
                'customer_po_name': po.name
            })
        return data

    def get(self, request, club_id):
        po_club = get_object_or_404(ActualPOClub, pk=club_id)
        po_club_quantity_data = self.get_po_club_pack_quantity_data(po_club)
        purchase_order_quantity_data = self.get_purchase_order_quantity_data(po_club)
        print(purchase_order_quantity_data)
        data = {
            "po_club_number": po_club.display_number,
            "style": po_club.pre_costing.order.style_number,
            "customer": po_club.pre_costing.order.customer.name,
            "order": po_club.pre_costing.order.short_code,
            "costing": po_club.pre_costing.short_code,
            "fabric_cost": [],
            "sewing_trim_cost": [],
            "packing_trim_cost": []
        }
        qs = GeneralPOMaterialQuantity.objects.filter(general_po__po_club=po_club).order_by(
            'material__material_detail__generic_material__user_material__display_order', 'material__id')
       
        for material_quantity in qs:
            serialize_data = GeneralPOMaterialQuantityListSerializer(material_quantity, many=False).data
            material_details = serialize_data["material_details"]
            size_display_value = next(
                (value for key, value in material_details.items() if key.endswith("_size_display_value")), None)
            if not size_display_value:
                size_display_value = ''
            
            supplier_inquiry_detail = material_quantity.default_material_supplier.supplier_inquiry_detail
            cost = self.calcuate_cost(supplier_inquiry_detail.supplier_inquiry.customer_brand_material, supplier_inquiry_detail.cost_per_unit, supplier_inquiry_detail.costing_unit, serialize_data['consumption'], 0)
            total_cost = self.calcuate_cost(supplier_inquiry_detail.supplier_inquiry.customer_brand_material, supplier_inquiry_detail.cost_per_unit, supplier_inquiry_detail.costing_unit, material_quantity.order_quantity, 0)
            
            material_data = {
                "material_label": material_quantity.material.material_label,
                "item_code": material_quantity.material.reference_code,
                "item_description": material_details['ritz_customer_brand_reference_code'],
                "size": size_display_value,
                "supplier": serialize_data['supplier_name'],
                "per_unit_qty": material_quantity.default_material_supplier.order_price,
                "consumption": round(serialize_data['consumption'], 5),
                "cost": round(cost, 5),
                "req_qty": serialize_data['quantity_display_value'],
                "wastage": serialize_data['wastage'],
                "qty_with_wastage": material_quantity.order_quantity,
                "grn_quantity": serialize_data['grn_quantity'],
                "balance_to_grn": serialize_data['variance'],
                "item_value": total_cost,
            }
            if serialize_data['material_category'] == Material.FABRIC_MATERIAL:
                data['fabric_cost'].append(material_data)
            elif serialize_data['material_category'] == SEWING_TRIM_TYPES:
                data['sewing_trim_cost'].append(material_data)
            elif serialize_data['material_category'] == PACKAGING_TYPES:
                data['packing_trim_cost'].append(material_data)
                
        response = self.create_bom_excel(po_club, data, purchase_order_quantity_data, po_club_quantity_data)
        return response


class GeneralPOSupplierPOListView(generics.ListAPIView):
    serializer_class = SupplierPOSerializer
    queryset = SupplierPO.objects.all().order_by('-id')
    permission_classes = (HasPermission,)
    write_roles = (MERCHANT_ROLE, )

    def get_queryset(self):
        id = self.kwargs.get('id', None)
        is_po_club = self.request.GET.get('is_po_club')
        if is_po_club == 'true':
            actual_club = get_object_or_404(ActualPOClub, pk=id)
            general_po_suppliers = GeneralPOSupplier.objects.filter(general_po__po_club=actual_club)
        else:
            general_po = get_object_or_404(GeneralPO, pk=id)
            general_po_suppliers = GeneralPOSupplier.objects.filter(general_po=general_po)
        qs = SupplierPO.objects.filter(general_po_supplier__in=general_po_suppliers)
        return qs
    

class GeneralPOFabricShadeSummary(generics.ListAPIView):
    permission_classes = (HasPermission, )
    write_roles = [MERCHANT_ROLE, ]
    serializer_class = GeneralPOShadeSummarySerializer

    def get_queryset(self):
        material_ids = FabricGRNDetail.objects.filter(
            grn_material_detail__supplier_po_grn_material__supplier_po_grn__supplier_po__general_po_supplier__general_po_id=self.kwargs['general_po_id']
        ).values_list(
            'grn_material_detail__supplier_po_grn_material__grn_material_id', flat=True
        )
        queryset = SupplierInquiryMaterialCode.objects.filter(id__in=material_ids).order_by('-id')
        return queryset

    def get_serializer_context(self):
        context = super().get_serializer_context()
        context['general_po_id'] = self.kwargs['general_po_id']
        return context
    
class GeneralPOFabricChartSummaryView(APIView):
    permission_classes = (HasPermission, )
    write_roles = [MERCHANT_ROLE, ]

    def get(self, request, general_po_id):
        response = {}
        material_ids = GeneralPOMaterialQuantity.objects.filter(
            general_po_id=general_po_id
        ).values_list('material', flat=True)
        materials = CustomerBrandMaterial.objects.filter(
            id__in=material_ids,
            material_code__material_definition__user_material__name=Material.FABRIC_MATERIAL
        )
        context = {'general_po_id': general_po_id}
        searializer = GeneralPOFabricChartSummarySerializer(materials, context=context, many=True).data
        return Response(searializer)
    

class GeneralPOMaterialQuantityDiscrepancyReasonsChoicesView(generics.ListAPIView):
    queryset = GeneralPOMaterialQuantity.DISCREPANCY_REASONS
    permission_classes = (HasPermission, )

    def list(self, request, *args, **kwargs):
        choices = [
            {"id": key, "name": value}
            for key, value in GeneralPOMaterialQuantity.DISCREPANCY_REASONS
        ]
        return Response(choices)


class GenerateSupplierPOsView(APIView):
    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]

    def post(self, request, *args, **kwargs):
        user = request.user
        data = {'success': True, 'message': 'Supplier POs generated successfully.'}
        return_status = status.HTTP_200_OK
        po_club_id = kwargs.get('po_club_id', None)
        po_club = get_object_or_404(ActualPOClub, pk=po_club_id)
        general_po_material_quantity_details = request.data.get('general_po_material_quantity_details', [])
        combine_supplier_po = request.data.get('combine_supplier_po', False)
        supplier_po_club_bom_generator = SupplierPOClubBOMGenerator(po_club)
        supplier_po_club_bom_generator.set_prepared_by(user)
        has_error, errors = supplier_po_club_bom_generator.validate_general_po_material_quantity_details(general_po_material_quantity_details)
        has_plant = supplier_po_club_bom_generator.all_purchase_order_has_plant()
        if has_error:
            data = errors
            return_status = status.HTTP_400_BAD_REQUEST
        elif not has_plant:
            data = {'plant': 'Please select a plant for all purchase orders'}
            return_status = status.HTTP_403_FORBIDDEN
        else:
            supplier_po_club_bom_generator.create_supplier_pos(general_po_material_quantity_details, combine_supplier_po)
        return Response(data=data, status=return_status)


class ReGenerateSupplierPOView(APIView):

    permission_classes = (HasPermission,)
    write_roles = [MERCHANT_ROLE, ]

    def post(self, request, *args, **kwargs):
        user = request.user
        data = {'status': 'Success'}
        errors = {}
        has_error = False
        return_status = status.HTTP_200_OK
        supplier_po_id = kwargs.get('supplier_po_id', None)
        supplier_po = get_object_or_404(SupplierPO, pk=supplier_po_id)
        po_club = supplier_po.get_po_club_or_none()
        if po_club:
            supplier_po.prepared_by = user
            supplier_po.save()
            supplier_po_bom_generator = SupplierPOClubBOMGenerator(po_club)
            supplier_po_bom_generator.set_prepared_by(user)
            supplier_po_bom_generator.re_generate_supplier_po(supplier_po)
        else:
            has_error = True
            errors = {
                'errors': ['Invalid PO Club']
            }
        if has_error:
            data = errors
            return_status = status.HTTP_400_BAD_REQUEST
        return Response(data=data, status=return_status)

