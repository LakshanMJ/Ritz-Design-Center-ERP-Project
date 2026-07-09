from openpyxl import load_workbook
from django.http import HttpResponse
from io import BytesIO
from marketing.utils.aws_utils import handle_file_read
from marketing.models import SupplierPOGRNMaterial, SupplierPOGRNMaterialDetail, FabricGRNDetail, FabricGRNBatchNumber, GRNBatchNumberShade
from rest_framework.response import Response
from rest_framework import status
import uuid
from materials.fieldmetadata.measuring_unit_helpers import MaterialUnitHelper
from shared.utils import get_object_or_none
from supplier_po.supplier_po_grn.serializers import MaterialGRNDetailSerializer, FabricGRNSerializer


class PackListProcessor:

    GENERAL_ERRORS_KEY = 'General Errors'

    _supplier_barcode_text = ['Supplier Barcode', ]
    _pack_no_text = ['Pack No', ]
    _batch_no_text = ['Batch No', ]
    _indicated_quantity_text = ['Indicated Quantity', ]
    _indicated_quantity_unit_text = ['Indicated Quantity Units', ]
    _indicated_width_text = ['Indicated Width',]
    _indicated_width_unit_text = ['Indicated Width Units', ]
    _gsm_text = ['Gsm', 'GSM']
    _wight_text = ['Weight', 'wight']
    _remarks_text = ['Remarks',]

    def __init__(self, grn_material, excel_file_path=None, attachment_id=None):
        super().__init__()
        self.grn_material = grn_material
        self.excel_file_path = excel_file_path
        self.attachment_id = attachment_id

    def download_pack_list_template(self):
        template_path = 'data_files/excel/material_pack_list_template.xlsx'
        wb = load_workbook(template_path)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        content_text = '%s%s%s%s' % ('attachment;', 'filename=', self.grn_material.id, '.xlsx')
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = content_text
        response.write(output.getvalue())
        wb.close()
        output.close()
        return response
    
    def get_batch_number_data(self, batch_number):
        data = {}
        value = None
        dispaly_value = batch_number
        fabric_batch_number = get_object_or_none(FabricGRNBatchNumber, {'grn_material': self.grn_material, 'batch_number': batch_number})
        if fabric_batch_number:
            value = fabric_batch_number.id
        data = {
            'value': value,
            'display_value': dispaly_value
        }
        return data
    
    def get_units_data(self, unit):
        data = {}
        for unit_value, display_value in MaterialUnitHelper.ALL_MEASURING_UNITS:
            if unit_value == unit:
                data = {
                    'value': unit_value,
                    'display_value': display_value
                }
                break
        return data
    
    def get_po_data(self, sheet, supplier_barcode_idx, pack_column_idx, batch_column_idx, quantity_column_idx, quantity_units_column_idx, width_column_idx, width_units_column_idx, gsm_column_idx, weight_column_idx, remarks_column_idx):
        grouped_data = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            if row[pack_column_idx]:
                batch_number = self.get_batch_number_data(row[batch_column_idx])
                quantity_unit = self.get_units_data(row[quantity_units_column_idx])
                width_unit = self.get_units_data(row[width_units_column_idx])
                grouped_data.append({
                    'id': None,
                    'supplier_po_grn_material': self.grn_material,
                    SupplierPOGRNMaterialDetail.BARCODE_VALUE_KEY: None,
                    SupplierPOGRNMaterialDetail.SUPPLIER_BARCODE_VALUE_KEY: row[supplier_barcode_idx],
                    FabricGRNDetail.PACK_NUMBER_VALUE_KEY: row[pack_column_idx],
                    SupplierPOGRNMaterialDetail.BATCH_NUMBER_VALUE_KEY: batch_number,
                    SupplierPOGRNMaterialDetail.INDICATED_QUANTITY_VALUE_KEY: row[quantity_column_idx],
                    SupplierPOGRNMaterialDetail.INDICATED_QUANTITY_UNITS_VALUE_KEY: quantity_unit,
                    FabricGRNDetail.INDICATED_WIDTH_VALUE_KEY: row[width_column_idx],
                    FabricGRNDetail.INDICATED_WIDTH_UNITS_VALUE_KEY: width_unit,
                    FabricGRNDetail.INDICATED_GSM_VALUE_KEY: row[gsm_column_idx],
                    FabricGRNDetail.INDICATED_WEIGHT_VALUE_KEY: row[weight_column_idx],
                    FabricGRNDetail.REMARKS_VALUE_KEY: row[remarks_column_idx],
                })
        return grouped_data

    def process_upload_pack_list(self):
        file = handle_file_read(self.excel_file_path)
        workbook = load_workbook(file)
        data = {}
        errors = None
        sheet = workbook.active
        errors = []
        have_errors, supplier_barcode_idx, pack_column_idx, batch_column_idx, \
        quantity_column_idx, quantity_units_column_idx, width_column_idx, width_units_column_idx, gsm_column_idx, weight_column_idx, remarks_column_idx = self.validate_and_get_header_column_index(sheet)

        # print(have_errors, material_column_idx, pack_column_idx, batch_column_idx, quantity_column_idx, width_column_idx, remarks_column_idx)

        if not have_errors:
            data = self.get_po_data(sheet, supplier_barcode_idx, pack_column_idx, batch_column_idx, quantity_column_idx, quantity_units_column_idx, width_column_idx, width_units_column_idx, gsm_column_idx, weight_column_idx, remarks_column_idx)
        else:
            errors.append({self.GENERAL_ERRORS_KEY :"Packing list format is not valid. Unable to locate headers. Please make sure the 1st row contains the headers."})
        workbook.close()
        response = self.create_material_detail(data)

        # print(response)

        return response

    def santize_value(self, value):
        return str(value).lower().strip()

    def validate_and_get_header_column_index(self, sheet):
        headers = [cell.value for cell in sheet[1]]
        header_indices = {header: idx for idx, header in enumerate(headers)}
        supplier_barcode_column_idx = None
        pack_column_idx = None
        batch_column_idx = None
        quantity_column_idx = None
        quantity_units_column_idx = None
        width_column_idx = None
        width_units_column_idx = None
        gsm_column_idx = None
        weight_column_idx = None
        remarks_column_idx = None

        for header in self._supplier_barcode_text:
            if header in header_indices:
                supplier_barcode_column_idx = header_indices[header]
                break
        for header in self._pack_no_text:
            if header in header_indices:
                pack_column_idx = header_indices[header]
                break
        for header in self._batch_no_text:
            if header in header_indices:
                batch_column_idx = header_indices[header]
                break
        for header in self._indicated_quantity_text:
            if header in header_indices:
                quantity_column_idx = header_indices[header]
                break
        for header in self._indicated_quantity_unit_text:
            if header in header_indices:
                quantity_units_column_idx = header_indices[header]
                break
        for header in self._indicated_width_text:
            if header in header_indices:
                width_column_idx = header_indices[header]
                break
        for header in self._indicated_width_unit_text:
            if header in header_indices:
                width_units_column_idx = header_indices[header]
                break
        for header in self._gsm_text:
            if header in header_indices:
                gsm_column_idx = header_indices[header]
                break
        for header in self._wight_text:
            if header in header_indices:
                weight_column_idx = header_indices[header]
                break
        for header in self._remarks_text:
            if header in header_indices:
                remarks_column_idx = header_indices[header]
                break
        if supplier_barcode_column_idx is None or pack_column_idx is None or batch_column_idx is None or quantity_column_idx is None or quantity_units_column_idx is None \
              or width_column_idx is None or width_units_column_idx is None or remarks_column_idx is None or weight_column_idx is None or remarks_column_idx is None:
            have_errors = True
        else:
            have_errors = False
        return have_errors, supplier_barcode_column_idx, pack_column_idx, batch_column_idx, quantity_column_idx, quantity_units_column_idx, width_column_idx, width_units_column_idx, gsm_column_idx, weight_column_idx, remarks_column_idx
    
    def get_barcode(self):
        lowercase_str = uuid.uuid4().hex
        return lowercase_str
    
    def create_material_detail(self, data):
        for row in data:
            print(row)
            grn_material = row.get('supplier_po_grn_material', None)
            fabric_batch_number, created = FabricGRNBatchNumber.objects.get_or_create(
                batch_number=row['batch_number']['display_value'],
                grn_material=grn_material
            )
            grn_material_detail = SupplierPOGRNMaterialDetail.objects.create(
                    supplier_po_grn_material=grn_material,
                    batch_number_id=fabric_batch_number.id,
                    supplier_barcode=row['supplier_barcode'],
                    indicated_quantity=row['indicated_quantity'],
                    indicated_quantity_units=row['indicated_quantity_units']['value'],
                    actual_quantity=row['indicated_quantity'],
                    actual_quantity_units=row['indicated_quantity_units']['value']
            )
            fabric_grn_detail = FabricGRNDetail.objects.get_or_create(
                grn_material_detail=grn_material_detail,
                pack_number=row['pack_number'],
                indicated_width = row['indicated_width'],
                indicated_width_units = row['indicated_width_units']['value'],
                indicated_gsm = row['indicated_gsm'],
                indicated_weight = row['indicated_weight'],
                actual_weight = row['indicated_weight'],
                remarks = row['remarks']
            )
            grn_material_detail.set_barcode()
            grn_material.set_material_quantities()
        return Response({'status': True})

